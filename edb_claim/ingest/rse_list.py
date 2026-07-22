"""Deterministic ingest of the ECMF-validated RSE list (PRD §4 input #2, FR-1).

This is the **authoritative roster of qualifying RSEs + citizenship/PR status**
(PRD §4) and the source of record for two eligibility gates:

  * **G1 — Local (SG citizen / PR)**: from the citizenship column here, cross-
    checked against the Time Sheet's Local/Foreigner column (PRD §6, FR-3).
  * **G2 — ECMF-validated RSE**: from the ``ecmf_validated`` flag here, cross-
    checked against the Time Sheet's ECMF flag (PRD §6, FR-3).

This is the **deterministic xlsx path** (openpyxl, read-only, no LLM). The
FR-9 LLM extraction path (T16) is a later augmentation that proposes the same
fields with confidence; it is NOT a dependency of this module.

Domain rules honoured (CLAUDE.md):
  * Foreigners and non-ECMF rows are **never silently dropped** — they are
    returned with their true status so the validate/ layer can EXCLUDE them
    *with reason* (PRD FR-3).
  * Every value carries an :class:`EvidenceRef` ``{file, sheet, cell/row}``
    (FR-7).
  * Read-only and deterministic: same file -> identical records, in sheet order
    (PRD §9).

Input schema is DEFINED here (the document is not in ``docs/``); the synthetic
generator (T14) must produce a workbook matching :data:`RSE_LIST_SCHEMA`.

----------------------------------------------------------------------------
Expected xlsx schema — sheet ``RSE List`` (header row 1, data from row 2)
----------------------------------------------------------------------------

| Column | Header              | Type   | Maps to                              |
|--------|---------------------|--------|--------------------------------------|
| A      | Employee ID         | str    | ``RseListRecord.employee_id`` (key)  |
| B      | Name                | str    | ``RseListRecord.name``               |
| C      | Citizenship         | str    | ``RseListRecord.citizenship`` (G1)   |
| D      | ECMF Validated      | bool   | ``RseListRecord.ecmf_validated`` (G2)|

  * **Citizenship** accepts the :class:`Citizenship` vocabulary plus common
    synonyms (case-insensitive): ``Citizen`` / ``Singapore Citizen`` / ``SC``,
    ``PR`` / ``Permanent Resident``, ``Foreigner`` / ``Foreign`` / ``EP`` /
    ``Work Pass``. Unrecognised values raise (fail loud — never guess local).
  * **ECMF Validated** accepts the booleans ``TRUE``/``FALSE``, ``Yes``/``No``,
    ``Y``/``N``, ``1``/``0`` (case-insensitive). Blank -> ``False`` (not
    validated until proven; conservative for G2).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Mapping, Optional, Tuple

from openpyxl import load_workbook

from edb_claim.domain.models import Citizenship, EvidenceRef


# --- DEFINED input schema (the contract T14 must generate to) --------------
DEFAULT_SHEET_NAME = "RSE List"
HEADER_ROW = 1
FIRST_DATA_ROW = 2

# Column letter -> logical field. Order is the audit/output order.
RSE_LIST_SCHEMA: Mapping[str, str] = {
    "A": "employee_id",
    "B": "name",
    "C": "citizenship",
    "D": "ecmf_validated",
}

# Citizenship synonym table -> canonical Citizenship enum (G1, PRD §6).
_CITIZENSHIP_SYNONYMS: Mapping[str, Citizenship] = {
    "citizen": Citizenship.CITIZEN,
    "singapore citizen": Citizenship.CITIZEN,
    "sg citizen": Citizenship.CITIZEN,
    "sc": Citizenship.CITIZEN,
    "pr": Citizenship.PR,
    "permanent resident": Citizenship.PR,
    "spr": Citizenship.PR,
    "foreigner": Citizenship.FOREIGNER,
    "foreign": Citizenship.FOREIGNER,
    "ep": Citizenship.FOREIGNER,
    "work pass": Citizenship.FOREIGNER,
}

# Boolean synonym table -> bool (ECMF flag, G2).
_TRUE_TOKENS = frozenset({"true", "yes", "y", "1", "validated", "ecmf"})
_FALSE_TOKENS = frozenset({"false", "no", "n", "0", "", "pending", "not validated"})


@dataclass(frozen=True)
class RseListRecord:
    """One row of the ECMF-validated RSE list (PRD §4 input #2).

    The authority for G1 (``citizenship``) and G2 (``ecmf_validated``). Carries
    an :class:`EvidenceRef` to its source row for FR-7 traceability. Frozen for
    the determinism guarantee (PRD §9) and safe reuse as a dict/set member.
    """

    employee_id: str
    name: str
    citizenship: Citizenship
    ecmf_validated: bool
    source_ref: Optional[EvidenceRef] = None


def _norm_cell(value: object) -> str:
    """Normalise a cell value to a stripped lower-case string ('' if blank)."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_citizenship(raw: object, locator: str) -> Citizenship:
    token = _norm_cell(raw)
    if token in _CITIZENSHIP_SYNONYMS:
        return _CITIZENSHIP_SYNONYMS[token]
    raise ValueError(
        f"Unrecognised citizenship {raw!r} at {locator}; expected one of "
        f"{sorted(set(_CITIZENSHIP_SYNONYMS))} (G1 must not be guessed)."
    )


def _parse_bool(raw: object, locator: str) -> bool:
    token = _norm_cell(raw)
    if token in _TRUE_TOKENS:
        return True
    if token in _FALSE_TOKENS:
        return False
    raise ValueError(
        f"Unrecognised ECMF-validated flag {raw!r} at {locator}; expected a "
        f"boolean-like token (e.g. TRUE/FALSE, Yes/No, Y/N, 1/0)."
    )


def parse_rse_list(
    path: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Tuple[RseListRecord, ...]:
    """Parse an ECMF-validated RSE list workbook into :class:`RseListRecord`s.

    Deterministic and read-only: returns one record per data row, in sheet
    order. Rows that are entirely blank are skipped (trailing-row tolerance);
    a row with a blank Employee ID but other content raises (fail loud — a
    nameless RSE cannot be keyed for G1/G2). Foreigners and non-ECMF rows are
    retained, never dropped (PRD FR-3).

    Args:
        path: filesystem path to the .xlsx workbook.
        sheet_name: worksheet holding the list (default ``"RSE List"``).

    Returns:
        Tuple of records in source-row order (determinism, PRD §9).

    Raises:
        KeyError: if ``sheet_name`` is absent.
        ValueError: on an unparseable citizenship/ECMF token or a partial row.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet {sheet_name!r} not found in {path!r}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        records: list[RseListRecord] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=FIRST_DATA_ROW), start=FIRST_DATA_ROW
        ):
            cells = {cell.column_letter: cell.value for cell in row}
            emp_id_raw = cells.get("A")
            name_raw = cells.get("B")
            cit_raw = cells.get("C")
            ecmf_raw = cells.get("D")

            # Fully blank row -> skip (tolerate trailing empties).
            if all(v is None or str(v).strip() == "" for v in (emp_id_raw, name_raw, cit_raw, ecmf_raw)):
                continue

            emp_id = "" if emp_id_raw is None else str(emp_id_raw).strip()
            if not emp_id:
                raise ValueError(
                    f"Blank Employee ID at {sheet_name}!A{row_idx} but the row "
                    f"has other content; cannot key this RSE for G1/G2."
                )

            records.append(
                RseListRecord(
                    employee_id=emp_id,
                    name="" if name_raw is None else str(name_raw).strip(),
                    citizenship=_parse_citizenship(cit_raw, f"{sheet_name}!C{row_idx}"),
                    ecmf_validated=_parse_bool(ecmf_raw, f"{sheet_name}!D{row_idx}"),
                    source_ref=EvidenceRef(
                        file=path,
                        sheet=sheet_name,
                        cell_or_row=f"A{row_idx}",
                        label="rse_list_row",
                    ),
                )
            )
        return tuple(records)
    finally:
        wb.close()


def index_by_employee_id(
    records: Tuple[RseListRecord, ...],
) -> Mapping[str, RseListRecord]:
    """Build an Employee-ID -> record lookup for G1/G2 cross-checks (FR-3).

    Raises on a duplicate Employee ID — the RSE list is the authority, so a
    duplicate is a data error that must surface, not be silently overwritten.
    """
    index: dict[str, RseListRecord] = {}
    for rec in records:
        if rec.employee_id in index:
            raise ValueError(
                f"Duplicate Employee ID {rec.employee_id!r} in RSE list; "
                f"the ECMF list must be unique per employee."
            )
        index[rec.employee_id] = rec
    return index
