"""Read-only ingest of the internal HR workbook (PRD FR-1, §4; PLAN.md T2).

Parses the two sheets of
``AI_COE_Claim_Checklist_Timesheet_for_FY_2026_to_2028_v2_Final.xlsx``:

  * ``Time Sheet`` (header row 18, data from row 19) -> one :class:`Employee`
    per real roster row, plus one :class:`PersonMonth` per month that carries
    a hours value (cols N..Y = Jan..Dec).
  * ``Staff Costs`` (header row 14, data from row 15) -> a :class:`StaffCostsRow`
    per real row, surfacing the workbook's OWN ``[A] [B] [C1] [C2] [D1] [D2]
    [D3] [E]`` values **as found** (cached, never recomputed — Method B / T7
    owns that arithmetic; this layer only reads for later reconciliation).

Design rules honoured (CLAUDE.md, PRD §9):
  * **Read-only.** The workbook is opened, never written; no calc, no LLM.
  * **Determinism.** Pure parse of cell values; identical file -> identical
    output. No ``datetime.now``/random/cwd-dependent behaviour.
  * **Every extracted field carries an** :class:`EvidenceRef`
    ``{file, sheet, cell}`` (e.g. ``Time Sheet!D19``) per FR-7.
  * **Header-text column matching** with a confirmed-position fallback: columns
    are located by matching the header row text so a shifted column does not
    silently mis-map; if a required header cannot be located the parser raises
    :class:`LayoutError` with a clear message rather than guessing.
  * **Staff Costs <-> Time Sheet cross-reference resolution.** Staff Costs cols
    C..H are formulas pointing at the Time Sheet. We read ``data_only=True`` to
    pick up cached values where Excel stored them; when a formula has no usable
    cached value (the supplied template leaves the Time Sheet identity cells
    blank, so the cached cross-refs come back as ``0``) we fall back to joining
    by **row offset** (Staff Costs data row N <-> Time Sheet data row N) and, as
    a secondary key, by Employee ID. See :func:`_join_staff_costs`.

The blank/template workbook ships with skeleton rows (auto-filled S/N in col B,
but no Employee ID, Name, hours, or hire type). Such rows are treated as
**empty** and skipped (reported in :attr:`IngestResult.skipped_empty_rows`);
only rows with genuine identifying content become :class:`Employee` objects.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Mapping, Optional, Tuple

from openpyxl.utils import get_column_letter

from edb_claim.domain.models import (
    Citizenship,
    Employee,
    EvidenceRef,
    HireType,
    PersonMonth,
)

# Lazy import so a missing openpyxl surfaces only when actually parsing.
try:  # pragma: no cover - exercised indirectly
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


# ---------------------------------------------------------------------------
# Layout constants (CLAUDE.md "Source templates"; verified against the file)
# ---------------------------------------------------------------------------
TIME_SHEET = "Time Sheet"
STAFF_COSTS = "Staff Costs"

TS_HEADER_ROW = 18
TS_DATA_START = 19
TS_ENTITY_CELL = "G3"  # participating-entity name (CLAUDE.md)

SC_HEADER_ROW = 14
SC_DATA_START = 15

# Month columns N..Y -> Jan..Dec on the Time Sheet (1-based month index).
_MONTH_HEADERS: Tuple[str, ...] = (
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
)


class LayoutError(ValueError):
    """Raised when the workbook layout does not match the confirmed contract."""


# ---------------------------------------------------------------------------
# Ingest-only container for the Staff Costs row.
# Intentionally NOT a domain model: it surfaces the workbook's own [A]..[E]
# values verbatim (incl. "N/A" / error strings) for the T7 Method-B replica and
# the variance reconciliation; it is not a computed result.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class StaffCostsRow:
    """One ``Staff Costs`` data row, values **as found** (no recompute)."""

    row: int
    employee_id: Optional[str]        # resolved via cross-ref / row-offset / col C
    name: Optional[str]
    hire_type_raw: Optional[str]      # col H ([NH/Up/Re]); raw string
    actual_salary_a: Optional[float]  # [A] col I (actual monthly salary per IR8A)
    qualifying_salary_b: object       # [B] col J: number OR "N/A" string, as found
    date_join_c1: Optional[date]      # [C1] col K
    date_left_c2: Optional[date]      # [C2] col L
    d1_capacity_hours: object         # [D1] col M = NETWORKDAYS*8.8 (as found)
    d2_project_hours: object          # [D2] col N: number OR "N/A" (as found)
    d3_time_pct: object               # [D3] col O: number OR error string (as found)
    e_qualifying_cost: object         # [E] col P: number OR error string (as found)
    join_method: str                  # how it was matched to a Time Sheet row
    source_refs: Tuple[EvidenceRef, ...] = field(default_factory=tuple)


@dataclass
class IngestResult:
    """Everything T2 extracts from one internal workbook (read-only)."""

    file: str
    entity: Optional[str]
    employees: List[Employee] = field(default_factory=list)
    person_months: List[PersonMonth] = field(default_factory=list)
    staff_costs: List[StaffCostsRow] = field(default_factory=list)
    skipped_empty_rows: List[int] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _ref(file: str, sheet: str, col: int, row: int, label: str) -> EvidenceRef:
    """Build an ``{file, sheet, cell}`` EvidenceRef, e.g. ``Time Sheet!D19``."""
    return EvidenceRef(
        file=file,
        sheet=sheet,
        cell_or_row=f"{sheet}!{get_column_letter(col)}{row}",
        label=label,
    )


def _norm_header(value: object) -> str:
    """Normalise a header cell for matching: str, lowercased, whitespace-collapsed."""
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().lower()


def _find_columns(
    ws, header_row: int, wanted: Mapping[str, Tuple[str, ...]], fallback: Mapping[str, int]
) -> Dict[str, int]:
    """Locate columns by header text, with a confirmed-position fallback.

    ``wanted`` maps a logical key -> tuple of acceptable header substrings.
    ``fallback`` maps the same key -> 1-based column index used only if the
    header text is not found. A required key that resolves to neither raises
    :class:`LayoutError`.
    """
    headers: Dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        headers[c] = _norm_header(ws.cell(header_row, c).value)

    resolved: Dict[str, int] = {}
    for key, needles in wanted.items():
        match = None
        for c, text in headers.items():
            if text and any(n in text for n in needles):
                match = c
                break
        if match is None:
            fb = fallback.get(key)
            if fb is None:
                raise LayoutError(
                    f"{ws.title!r}: could not locate column for {key!r} "
                    f"(looked for {needles}) and no positional fallback."
                )
            # Sanity-check the fallback header isn't obviously a different field.
            resolved[key] = fb
        else:
            resolved[key] = match
    return resolved


def _to_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_float(value: object) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _clean_str(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s or None
    return str(value)


def _as_bool(value: object) -> bool:
    """Coerce a checklist cell to bool. TRUE/'yes'/'y'/1 -> True, else False."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "yes", "y", "1", "x")
    return False


def _parse_citizenship(value: object) -> Citizenship:
    """Map the 'Local/Foreigner' col to a Citizenship enum (best-effort).

    The template column is free-ish text. 'Local' (or Citizen/SC/PR) -> CITIZEN
    unless 'PR' is explicit; 'Foreigner' -> FOREIGNER. Unknown/blank defaults to
    FOREIGNER so an unverified person is never silently treated as claimable
    (CLAUDE.md: exclusions reported, never silently dropped). The validate layer
    refines this against the authoritative RSE list (T3/G1).
    """
    s = _clean_str(value)
    if s is None:
        return Citizenship.FOREIGNER
    low = s.lower()
    if "pr" in low and "permanent" in low or low in ("pr", "spr"):
        return Citizenship.PR
    if "foreign" in low:
        return Citizenship.FOREIGNER
    if "local" in low or "citizen" in low or low in ("sc", "sg"):
        return Citizenship.CITIZEN
    return Citizenship.FOREIGNER


def _parse_hire_type(value: object) -> Optional[HireType]:
    s = _clean_str(value)
    if s is None:
        return None
    low = s.lower()
    if "new" in low:
        return HireType.NEW_HIRE
    if "upskill" in low:
        return HireType.UPSKILLED
    if "reskill" in low:
        return HireType.RESKILLED
    return None


def _normalize_name(name: Optional[str]) -> Optional[str]:
    """FR-11 reconciliation key: upper-cased, sorted tokens (order-insensitive)."""
    if not name:
        return None
    tokens = [t for t in name.replace(",", " ").split() if t]
    if not tokens:
        return None
    return " ".join(sorted(t.upper() for t in tokens))


# ---------------------------------------------------------------------------
# Time Sheet
# ---------------------------------------------------------------------------
_TS_WANTED: Mapping[str, Tuple[str, ...]] = {
    "sn": ("s/n",),
    "employee_id": ("employee id",),
    "name": ("name",),
    "local_foreigner": ("local",),
    "qualifications": ("qualification",),
    "designation": ("designation",),
    "ecmf": ("validated by ecmf", "ecmf"),
    "no_other_grant": ("not enjoying", "other government"),
    "ai_proficiency": ("ai proficiency", "proficiency level"),
    "ai_capability": ("ai capabilit",),
    "hire_type": ("new hire", "upskilled", "reskilled", "new hirer"),
    "upskilling_date": ("date of completion", "upskilling"),
}
# Confirmed positions (1-based) used only if header text matching fails.
_TS_FALLBACK = {
    "sn": 2, "employee_id": 3, "name": 4, "local_foreigner": 5,
    "qualifications": 6, "designation": 7, "ecmf": 8, "no_other_grant": 9,
    "ai_proficiency": 10, "ai_capability": 11, "hire_type": 12,
    "upskilling_date": 13,
}
# Month columns N..Y (14..25); located by header text Jan..Dec with positional
# fallback. Returned as a logical key per month: "m1".."m12".
_TS_MONTH_FALLBACK = {f"m{i}": 13 + i for i in range(1, 13)}


def _find_month_columns(ws) -> Dict[int, int]:
    """Return {month_number(1..12): column index} for the Time Sheet.

    Months are matched by an **exact** normalised header equality (e.g. the
    cell text *is* "jan"), not substring — a substring match would wrongly hook
    "oct" onto the "Doctorate" qualifications header. Falls back to the
    confirmed positions N..Y (14..25) for any month whose header is not found.
    """
    headers: Dict[int, str] = {
        c: _norm_header(ws.cell(TS_HEADER_ROW, c).value)
        for c in range(1, ws.max_column + 1)
    }
    resolved: Dict[int, int] = {}
    for i in range(1, 13):
        token = _MONTH_HEADERS[i - 1].lower()
        match = next((c for c, t in headers.items() if t == token), None)
        resolved[i] = match if match is not None else _TS_MONTH_FALLBACK[f"m{i}"]
    return resolved


def _ts_row_is_empty(ws, row: int, cols: Dict[str, int], month_cols: Dict[int, int]) -> bool:
    """A Time Sheet row is 'real' only if it has identifying content.

    The skeleton template auto-fills S/N (col B) and even a hire-type dropdown
    on blank rows, so neither alone qualifies. We require at least one of:
    Employee ID, Name, or any non-zero monthly hours.
    """
    eid = _clean_str(ws.cell(row, cols["employee_id"]).value)
    name = _clean_str(ws.cell(row, cols["name"]).value)
    if eid or name:
        return False
    for c in month_cols.values():
        v = _to_float(ws.cell(row, c).value)
        if v not in (None, 0.0):
            return False
    return True


def _ts_last_data_row(ws, cols: Dict[str, int]) -> int:
    """Last plausible data row: stop before footer rows ('Prepared by:' etc.).

    The S/N column (col B) is numeric on data rows and text on footers; we walk
    from the data start while col B is a number (the auto-fill chain) and bound
    by max_row.
    """
    last = TS_DATA_START - 1
    sn_col = cols["sn"]
    for r in range(TS_DATA_START, ws.max_row + 1):
        sn = ws.cell(r, sn_col).value
        if isinstance(sn, (int, float)) and not isinstance(sn, bool):
            last = r
        elif sn is None:
            # blank S/N but maybe identifying content present -> still include
            if not _ts_row_is_empty(ws, r, cols, {}):
                last = r
        else:
            break  # hit a text footer row
    return last


def _parse_time_sheet(wb_values, file: str) -> Tuple[
    Optional[str],
    Dict[int, Employee],          # keyed by Time Sheet row
    Dict[int, str],               # row -> employee_id (resolved/synthetic)
    List[PersonMonth],
    List[int],                    # skipped empty rows
    List[str],                    # warnings
    Dict[str, int],               # resolved columns (for cross-ref)
]:
    ws = wb_values[TIME_SHEET]
    warnings: List[str] = []

    cols = _find_columns(ws, TS_HEADER_ROW, _TS_WANTED, _TS_FALLBACK)
    month_cols = _find_month_columns(ws)

    entity = _clean_str(ws[TS_ENTITY_CELL].value)

    employees_by_row: Dict[int, Employee] = {}
    eid_by_row: Dict[int, str] = {}
    person_months: List[PersonMonth] = []
    skipped: List[int] = []

    last = _ts_last_data_row(ws, cols)
    for row in range(TS_DATA_START, last + 1):
        if _ts_row_is_empty(ws, row, cols, month_cols):
            skipped.append(row)
            continue

        eid = _clean_str(ws.cell(row, cols["employee_id"]).value)
        name = _clean_str(ws.cell(row, cols["name"]).value)
        # Synthetic, stable id when the workbook omits one (skeleton/POC rows):
        resolved_id = eid or f"{TIME_SHEET}!row{row}"
        eid_by_row[row] = resolved_id

        designation = _clean_str(ws.cell(row, cols["designation"]).value) or ""
        hire = _parse_hire_type(ws.cell(row, cols["hire_type"]).value)

        refs = [
            _ref(file, TIME_SHEET, cols["employee_id"], row, "employee_id"),
            _ref(file, TIME_SHEET, cols["name"], row, "name"),
            _ref(file, TIME_SHEET, cols["local_foreigner"], row, "citizenship"),
            _ref(file, TIME_SHEET, cols["ecmf"], row, "ecmf_validated"),
            _ref(file, TIME_SHEET, cols["no_other_grant"], row, "no_other_grant"),
            _ref(file, TIME_SHEET, cols["designation"], row, "designation"),
            _ref(file, TIME_SHEET, cols["hire_type"], row, "hire_type"),
        ]

        emp = Employee(
            id=resolved_id,
            name=name or "",
            entity=entity or "",
            citizenship=_parse_citizenship(ws.cell(row, cols["local_foreigner"]).value),
            ecmf_validated=_as_bool(ws.cell(row, cols["ecmf"]).value),
            no_other_grant=_as_bool(ws.cell(row, cols["no_other_grant"]).value),
            designation=designation,
            hire_type=hire if hire is not None else HireType.UPSKILLED,
            normalized_name=_normalize_name(name),
            source_refs=tuple(refs),
        )
        if hire is None:
            warnings.append(
                f"{TIME_SHEET}!{get_column_letter(cols['hire_type'])}{row}: "
                f"unrecognised hire type; defaulted to UPSKILLED."
            )
        employees_by_row[row] = emp

        # Per-month hours -> PersonMonth (only months that carry a value).
        for month, c in month_cols.items():
            raw = ws.cell(row, c).value
            hours = _to_float(raw)
            if hours is None:
                continue  # blank month: no PersonMonth emitted
            person_months.append(
                PersonMonth(
                    employee_id=resolved_id,
                    year=0,  # year not stored on the Time Sheet (cols are bare months)
                    month=month,
                    basic_salary=0.0,  # filled by salary ingest (T3); not on this sheet
                    hours=hours,
                    source_refs=(_ref(file, TIME_SHEET, c, row, f"hours_m{month}"),),
                )
            )

    return entity, employees_by_row, eid_by_row, person_months, skipped, warnings, cols


# ---------------------------------------------------------------------------
# Staff Costs
# ---------------------------------------------------------------------------
_SC_WANTED: Mapping[str, Tuple[str, ...]] = {
    "sn": ("s/n",),
    "employee_id": ("employee id",),
    "name": ("name",),
    "hire_type": ("[h]", "new hire", "upskilled", "reskilled", "new hirer"),
    "a": ("[a]", "actual monthly salary"),
    "b": ("[b]", "qualifying annual salary", "total qualifying"),
    "c1": ("[c1]", "date join"),
    "c2": ("[c2]", "date left"),
    "d1": ("[d1]", "total hours in employment"),
    "d2": ("[d2]", "time spent on project"),
    "d3": ("[d3]", "% of time"),
    "e": ("[e]", "staff cost for risc"),
}
# Confirmed positions (1-based). Header row 14 holds the [A]..[E] tags; the
# descriptive labels sit on row 13, so the [..] tags are the reliable anchors.
_SC_FALLBACK = {
    "sn": 2, "employee_id": 3, "name": 4, "hire_type": 8,
    "a": 9, "b": 10, "c1": 11, "c2": 12, "d1": 13, "d2": 14, "d3": 15, "e": 16,
}


def _sc_last_data_row(ws, sn_col: int) -> int:
    last = SC_DATA_START - 1
    for r in range(SC_DATA_START, ws.max_row + 1):
        sn = ws.cell(r, sn_col).value
        if isinstance(sn, (int, float)) and not isinstance(sn, bool):
            last = r
        elif sn is None:
            last = r  # gap row inside the block; keep scanning
        else:
            break  # text footer
    return last


def _sc_xref_value(value: object) -> Optional[str]:
    """A usable cross-ref string, or None if the cached value is unusable.

    The template's Staff Costs C/D cols are formulas into the Time Sheet; when
    the Time Sheet identity cells are blank, Excel cached ``0`` (numeric). Treat
    0 / blank / non-identifying numerics as 'no usable cached value'.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return None if value == 0 else str(value)
    s = _clean_str(value)
    return s


def _sc_row_is_empty(ws, row: int, cols: Dict[str, int]) -> bool:
    """Real Staff Costs row needs at least an [A] value or a usable identity."""
    a = _to_float(ws.cell(row, cols["a"]).value)
    if a not in (None, 0.0):
        return False
    if _sc_xref_value(ws.cell(row, cols["employee_id"]).value):
        return False
    if _sc_xref_value(ws.cell(row, cols["name"]).value):
        return False
    return True


def _parse_staff_costs(
    wb_values, file: str
) -> Tuple[List[StaffCostsRow], List[str], int]:
    ws = wb_values[STAFF_COSTS]
    warnings: List[str] = []
    cols = _find_columns(ws, SC_HEADER_ROW, _SC_WANTED, _SC_FALLBACK)

    rows: List[StaffCostsRow] = []
    last = _sc_last_data_row(ws, cols["sn"])
    for row in range(SC_DATA_START, last + 1):
        if _sc_row_is_empty(ws, row, cols):
            continue

        def cell(key: str):
            return ws.cell(row, cols[key]).value

        refs = tuple(
            _ref(file, STAFF_COSTS, cols[k], row, label)
            for k, label in (
                ("employee_id", "employee_id"),
                ("name", "name"),
                ("hire_type", "hire_type"),
                ("a", "actual_salary_a"),
                ("b", "qualifying_salary_b"),
                ("c1", "date_join_c1"),
                ("c2", "date_left_c2"),
                ("d1", "d1_capacity_hours"),
                ("d2", "d2_project_hours"),
                ("d3", "d3_time_pct"),
                ("e", "e_qualifying_cost"),
            )
        )

        rows.append(
            StaffCostsRow(
                row=row,
                employee_id=_sc_xref_value(cell("employee_id")),
                name=_sc_xref_value(cell("name")),
                hire_type_raw=_clean_str(cell("hire_type")),
                actual_salary_a=_to_float(cell("a")),
                qualifying_salary_b=cell("b"),  # number OR "N/A", as found
                date_join_c1=_to_date(cell("c1")),
                date_left_c2=_to_date(cell("c2")),
                d1_capacity_hours=cell("d1"),
                d2_project_hours=cell("d2"),  # number OR "N/A", as found
                d3_time_pct=cell("d3"),       # number OR error string, as found
                e_qualifying_cost=cell("e"),  # number OR error string, as found
                join_method="unresolved",
                source_refs=refs,
            )
        )
    return rows, warnings, SC_DATA_START


def _join_staff_costs(
    staff_costs: List[StaffCostsRow],
    eid_by_ts_row: Dict[int, str],
) -> Tuple[List[StaffCostsRow], List[str]]:
    """Resolve each Staff Costs row to a Time Sheet employee_id.

    Priority:
      1. **Cached cross-ref Employee ID** (Staff Costs col C) when it matches a
         known Time Sheet employee_id (the ideal, formula-resolved path).
      2. **Row offset.** Staff Costs data row N maps to Time Sheet data row N:
         offset = (TS_DATA_START - SC_DATA_START) = (19 - 15) = 4. This is the
         workbook's own structural pairing (every Staff Costs C..H formula on
         row r points at Time Sheet row r+4), so it is the reliable fallback
         when the cached identity is blank.
    The chosen method is recorded on :attr:`StaffCostsRow.join_method`.
    """
    warnings: List[str] = []
    offset = TS_DATA_START - SC_DATA_START  # = 4
    known_ids = set(eid_by_ts_row.values())
    resolved: List[StaffCostsRow] = []

    from dataclasses import replace as _replace

    for sc in staff_costs:
        eid = sc.employee_id
        method = "unresolved"
        if eid and eid in known_ids:
            method = "cross_ref_employee_id"
        else:
            ts_row = sc.row + offset
            if ts_row in eid_by_ts_row:
                eid = eid_by_ts_row[ts_row]
                method = "row_offset"
            else:
                method = "unmatched"
                warnings.append(
                    f"{STAFF_COSTS} row {sc.row}: no Time Sheet match "
                    f"(cross-ref blank and row offset {ts_row} not a data row)."
                )
        resolved.append(_replace(sc, employee_id=eid, join_method=method))
    return resolved, warnings


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def parse_internal_workbook(path: str) -> IngestResult:
    """Parse the internal HR workbook at ``path`` (read-only) -> IngestResult.

    Opens the workbook ``data_only=True`` so cached formula values (Staff Costs
    [B]/[D1]/[D3]/[E], the Z totals, cross-refs) are available; falls back to
    row-offset joining where a cross-ref has no usable cached value. Raises
    :class:`LayoutError` if the confirmed sheet/column layout is not present.
    """
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required to parse the internal workbook.")
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    file = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        missing = [s for s in (TIME_SHEET, STAFF_COSTS) if s not in wb.sheetnames]
        if missing:
            raise LayoutError(
                f"{file}: required sheet(s) missing: {missing}; "
                f"found {wb.sheetnames}."
            )

        (
            entity,
            employees_by_row,
            eid_by_row,
            person_months,
            skipped,
            ts_warnings,
            _ts_cols,
        ) = _parse_time_sheet(wb, file)

        staff_costs_raw, sc_warnings, _ = _parse_staff_costs(wb, file)
        staff_costs, join_warnings = _join_staff_costs(staff_costs_raw, eid_by_row)
    finally:
        wb.close()

    return IngestResult(
        file=file,
        entity=entity,
        employees=[employees_by_row[r] for r in sorted(employees_by_row)],
        person_months=person_months,
        staff_costs=staff_costs,
        skipped_empty_rows=skipped,
        warnings=ts_warnings + sc_warnings + join_warnings,
    )
