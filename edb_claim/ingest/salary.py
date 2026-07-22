"""Deterministic ingest of the payslip / payroll register (PRD §4 input #3, FR-1).

Produces one :class:`SalaryRecord` per **employee-month**, carrying the
**basic monthly salary ONLY** (PRD §6, CLAUDE.md). This is the salary evidence
for two layers downstream:

  * **G4 floor** — basic monthly salary < S$5,000 -> person-month EXCLUDED
    (gate applied in validate/, not here).
  * **S$20,000 cap** — arithmetic clamp applied per month in calc/ (not here).

This is the **deterministic xlsx path** (openpyxl, read-only, no LLM). The
FR-9 LLM extraction path (T16) — which also captures allowance/bonus/CPF
components precisely to *exclude* them — is a later augmentation, NOT a
dependency of this module.

Domain rule — what "basic salary" EXCLUDES and why
--------------------------------------------------
"Qualifying salary = basic monthly salary only" (PRD §6, CLAUDE.md). The
following columns, if present, are **deliberately ignored** and never summed
into ``basic_salary``:

  * **CPF** (employer + employee contributions) — statutory, not salary.
  * **Bonus** / variable / performance pay — not basic.
  * **AWS** (13th-month / Annual Wage Supplement) — not basic.
  * **Allowances** — transport, housing, meal, shift, etc.
  * **COLA** (cost-of-living adjustment) — not basic.
  * **Airfare** / relocation — not basic.
  * **OT** (overtime) — not basic.
  * **Gross / Net / Total Pay** — aggregates that *include* the excluded
    components; reading these would over-claim. We read **Basic Salary** only.

These exclusions matter because using gross/total (or adding CPF/bonus) would
inflate the qualifying cost and the claim — a hard audit failure under SSRS
4400. The deterministic path therefore reads exactly one designated column
(Basic Salary) and treats every other money column as informational.

Input schema is DEFINED here (the document is not in ``docs/``); the synthetic
generator (T14) must produce a workbook matching :data:`PAYROLL_SCHEMA`.

----------------------------------------------------------------------------
Expected xlsx schema — sheet ``Payroll`` (header row 1, data from row 2)
----------------------------------------------------------------------------
**Long / tidy** layout: one row per employee-month.

| Column | Header         | Type  | Used? | Maps to / note                    |
|--------|----------------|-------|-------|-----------------------------------|
| A      | Employee ID    | str   | yes   | ``SalaryRecord.employee_id``      |
| B      | Name           | str   | info  | not stored (RSE list is authority)|
| C      | Year           | int   | yes   | ``SalaryRecord.year``             |
| D      | Month          | int   | yes   | ``SalaryRecord.month`` (1-12)     |
| E      | Basic Salary   | float | yes   | ``SalaryRecord.basic_salary``     |
| F      | Allowances     | float | NO    | EXCLUDED (see above)              |
| G      | Bonus          | float | NO    | EXCLUDED                          |
| H      | AWS            | float | NO    | EXCLUDED                          |
| I      | CPF (Employer) | float | NO    | EXCLUDED                          |
| J      | CPF (Employee) | float | NO    | EXCLUDED                          |
| K      | Gross Pay      | float | NO    | EXCLUDED (aggregate)              |

  * Required columns: ``Employee ID``, ``Year``, ``Month``, ``Basic Salary``.
    Any of the excluded columns may be present or absent — both are fine; they
    are never read into the figure.
  * ``Month`` is the numeric month 1-12 (``Year``-``Month`` is the period key).
  * Columns are located **by header text** (case-insensitive, synonym-aware),
    not by fixed letter, so the register can carry extra columns in any order —
    but the four required headers must be findable.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from typing import Mapping, Optional, Tuple

from openpyxl import load_workbook

from edb_claim.domain.models import EvidenceRef, SalaryRecord


# --- DEFINED input schema (the contract T14 must generate to) --------------
DEFAULT_SHEET_NAME = "Payroll"
HEADER_ROW = 1
FIRST_DATA_ROW = 2

# Header synonyms (exact, lower-cased) -> logical field. Backed by a contains
# fallback (_FIELD_CONTAINS) and a combined Period/Pay-Date column, so a wide
# range of real payroll layouts parse without manual mapping.
_HEADER_SYNONYMS: Mapping[str, str] = {
    "employee id": "employee_id", "employee_id": "employee_id", "emp id": "employee_id",
    "emp no": "employee_id", "employee no": "employee_id", "staff id": "employee_id",
    "staff no": "employee_id", "payroll id": "employee_id", "employee code": "employee_id",
    "staff code": "employee_id", "emp code": "employee_id", "emp. no": "employee_id",
    "name": "name", "employee name": "name", "staff name": "name", "full name": "name",
    "year": "year", "yr": "year", "fy": "year",
    "month": "month", "mth": "month", "mon": "month",
    "period": "period", "pay period": "period", "pay date": "period", "payment date": "period",
    "month-year": "period", "month year": "period", "pay month": "period", "salary month": "period",
    "pay-date": "period", "date": "period",
    "basic salary": "basic_salary", "basic": "basic_salary", "basic monthly salary": "basic_salary",
    "basic pay": "basic_salary", "base salary": "basic_salary", "base pay": "basic_salary",
    "basic wage": "basic_salary", "monthly basic": "basic_salary", "monthly salary": "basic_salary",
    "salary": "basic_salary", "basic salary (sgd)": "basic_salary",
}

# Ordered contains-fallback when no exact synonym matches. First field whose
# keyword is a substring of the header wins (priority high -> low).
_FIELD_CONTAINS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("employee_id", ("employee id", "employee no", "emp id", "emp no", "staff id",
                     "staff no", "payroll id", "employee code", "staff code", "emp code")),
    ("name", ("employee name", "staff name", "full name", "name")),
    ("period", ("pay period", "pay date", "payment date", "month-year", "month year",
                "pay month", "salary month", "period")),
    ("year", ("year", "fy")),
    ("month", ("month", "mth")),
    ("basic_salary", ("basic monthly salary", "basic salary", "basic pay", "base salary",
                      "base pay", "basic wage", "monthly basic", "monthly salary", "basic", "salary")),
)

# Only the salary column is mandatory; employee id and the period (year/month)
# can be inferred from a Period column or the file name (single-payslip files).
REQUIRED_FIELDS = ("basic_salary",)

_MONTH_NAMES: Mapping[str, int] = {
    name: i
    for i, names in enumerate(
        (("jan", "january"), ("feb", "february"), ("mar", "march"), ("apr", "april"),
         ("may",), ("jun", "june"), ("jul", "july"), ("aug", "august"),
         ("sep", "sept", "september"), ("oct", "october"), ("nov", "november"),
         ("dec", "december")),
        start=1)
    for name in names
}

# Documented exclusion list (informational; the module's contract). These are
# never mapped to a logical field, so they can never feed ``basic_salary``.
EXCLUDED_COMPONENTS: Tuple[str, ...] = (
    "cpf",
    "bonus",
    "aws",
    "allowance",
    "cola",
    "airfare",
    "relocation",
    "overtime",
    "ot",
    "gross",
    "net",
    "total pay",
    "total",
)


class PayrollSchemaError(Exception):
    """Raised when required headers are missing/ambiguous in the register."""


def _norm(value: object) -> str:
    return "" if value is None else str(value).strip().lower()


def _classify_header(h: str) -> Optional[str]:
    """Map a normalised header to a logical field, or ``None``.

    Excluded components (CPF/bonus/AWS/gross/net/…) are never mapped, so a
    'Gross Salary' or 'CPF' column can't be mistaken for basic salary. Tries an
    exact synonym first, then an ordered substring fallback (so 'Basic Salary
    (SGD)' or 'Pay Month' still resolve).
    """
    if any(x in h for x in EXCLUDED_COMPONENTS):
        return None
    exact = _HEADER_SYNONYMS.get(h)
    if exact is not None:
        return exact
    for field, kws in _FIELD_CONTAINS:
        if any(kw in h for kw in kws):
            return field
    return None


def _map_headers(header_cells, sheet_name: str) -> Mapping[str, str]:
    """Map logical fields -> column letters by header text (synonym + contains).

    Only the basic-salary column is mandatory; employee id and the period
    (year/month, or a combined Period column) are optional here and may be
    inferred from the file name for single-payslip files.

    Raises:
        PayrollSchemaError: no basic-salary column can be found.
    """
    field_to_col: dict[str, str] = {}
    for cell in header_cells:
        header = _norm(cell.value)
        if not header:
            continue
        logical = _classify_header(header)
        if logical is None or logical in field_to_col:
            continue  # first match wins; tolerate duplicates/extra columns
        field_to_col[logical] = cell.column_letter

    if "basic_salary" not in field_to_col:
        raise PayrollSchemaError(
            f"Payroll sheet {sheet_name!r}: couldn't find a basic-salary column "
            f"(looked for e.g. 'Basic Salary' / 'Salary'). Found: {field_to_col}."
        )
    return field_to_col


def _to_int(raw: object, locator: str, what: str) -> int:
    if raw is None or str(raw).strip() == "":
        raise ValueError(f"Missing {what} at {locator}.")
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError):
        raise ValueError(f"Non-integer {what} {raw!r} at {locator}.")


def _to_float(raw: object, locator: str, what: str) -> float:
    if raw is None or str(raw).strip() == "":
        raise ValueError(f"Missing {what} at {locator}.")
    try:
        # Tolerate thousands separators / currency symbols in synthetic data.
        cleaned = str(raw).replace(",", "").replace("$", "").strip()
        return float(cleaned)
    except (TypeError, ValueError):
        raise ValueError(f"Non-numeric {what} {raw!r} at {locator}.")


# --- tolerant period / id parsing (columns, date cells, or the file name) ---
_RE_YM_SEP = re.compile(r"((?:19|20)\d{2})[\-_/.\s](0?[1-9]|1[0-2])(?!\d)")   # 2024-01, 2024/1
_RE_YM_COMPACT = re.compile(r"((?:19|20)\d{2})(0[1-9]|1[0-2])(?!\d)")          # 202401
_RE_MY_SEP = re.compile(r"(0?[1-9]|1[0-2])[\-_/.\s]((?:19|20)\d{2})")          # 01-2024
_RE_MONTH_NAME = re.compile(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*", re.I)
_RE_YEAR = re.compile(r"(19|20)\d{2}")
_RE_EMP_ID = re.compile(r"\b([A-Za-z]{1,6}-?\d{2,6})\b")


def _infer_period_from_text(s: str) -> Optional[Tuple[int, int]]:
    """Pull a (year, month) out of free text: 'Jan 2024', '2024-01', '202401'…"""
    mn = _RE_MONTH_NAME.search(s)
    yr = _RE_YEAR.search(s)
    if mn and yr:
        mon = _MONTH_NAMES.get(mn.group(0).lower()[:3])
        if mon:
            return int(yr.group(0)), mon
    for rx, yi, mi in ((_RE_YM_SEP, 1, 2), (_RE_YM_COMPACT, 1, 2), (_RE_MY_SEP, 2, 1)):
        m = rx.search(s)
        if m:
            return int(m.group(yi)), int(m.group(mi))
    return None


def _infer_period_from_name(name: str) -> Optional[Tuple[int, int]]:
    return _infer_period_from_text(os.path.splitext(os.path.basename(name))[0])


def _infer_employee_id_from_name(name: str) -> Optional[str]:
    base = os.path.splitext(os.path.basename(name))[0]
    m = _RE_EMP_ID.search(base)
    return m.group(1).upper() if m else None


def _parse_year(raw: object, locator: str) -> int:
    if isinstance(raw, (datetime, date)):
        return raw.year
    m = _RE_YEAR.search(str(raw))
    if not m:
        raise ValueError(f"Couldn't read a year from {raw!r} at {locator}.")
    return int(m.group(0))


def _parse_month(raw: object, locator: str) -> int:
    if isinstance(raw, (datetime, date)):
        return raw.month
    s = str(raw).strip().lower()
    if s.replace(".0", "").isdigit():
        mth = int(float(s))
    else:
        mn = _RE_MONTH_NAME.match(s) or _RE_MONTH_NAME.search(s)
        if mn:
            mth = _MONTH_NAMES.get(mn.group(0)[:3], 0)
        else:
            num = re.search(r"\b(0?[1-9]|1[0-2])\b", s)
            mth = int(num.group(1)) if num else 0
    if not 1 <= mth <= 12:
        raise ValueError(f"Couldn't read a valid month (1-12) from {raw!r} at {locator}.")
    return mth


def _parse_period_value(raw: object, locator: str) -> Tuple[int, int]:
    if isinstance(raw, (datetime, date)):
        return raw.year, raw.month
    p = _infer_period_from_text(str(raw))
    if p is None:
        raise ValueError(f"Couldn't read a year/month period from {raw!r} at {locator}.")
    return p


_PAYROLL_SHEET_ALIASES = ("payroll", "payslip", "payslips", "salary", "salaries", "pay register", "wages")


def _resolve_payroll_sheet(wb, requested: str) -> Optional[str]:
    """Pick the payroll worksheet tolerantly: exact → case-insensitive → alias →
    single-sheet fallback. Returns the actual sheet name, or ``None`` if unresolved.

    Lets HR upload a register whose tab is named e.g. ``Payslip`` instead of the
    canonical ``Payroll`` without a hard failure.
    """
    names = list(wb.sheetnames)
    if requested in names:
        return requested
    low = {n.lower(): n for n in names}
    if requested.lower() in low:
        return low[requested.lower()]
    for n in names:  # alias contained in the tab name
        if any(a in n.lower() for a in _PAYROLL_SHEET_ALIASES):
            return n
    if len(names) == 1:  # a single-sheet workbook is unambiguous
        return names[0]
    return None


def parse_payroll_register(
    path: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Tuple[SalaryRecord, ...]:
    """Parse a payslip/payroll register into per-employee-month basic salary.

    Reads **basic monthly salary only**; every CPF/bonus/AWS/allowance/gross
    column is ignored (see module docstring). Deterministic and read-only:
    one :class:`SalaryRecord` per data row, in sheet order.

    Args:
        path: filesystem path to the .xlsx register.
        sheet_name: worksheet holding the register (default ``"Payroll"``).

    Returns:
        Tuple of :class:`SalaryRecord`, source-row order (determinism, PRD §9).
        Each carries an :class:`EvidenceRef` to its Basic-Salary cell (FR-7).

    Raises:
        KeyError: if ``sheet_name`` is absent.
        PayrollSchemaError: if a required column header is missing.
        ValueError: on an unparseable year/month/basic in a non-blank row, or a
            duplicate (employee_id, year, month) period.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        resolved = _resolve_payroll_sheet(wb, sheet_name)
        if resolved is None:
            raise KeyError(
                f"Sheet {sheet_name!r} not found in {path!r}; "
                f"available: {wb.sheetnames}"
            )
        sheet_name = resolved
        ws = wb[sheet_name]

        header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))
        field_to_col = _map_headers(header_row, sheet_name)
        col_id = field_to_col.get("employee_id")
        col_year = field_to_col.get("year")
        col_month = field_to_col.get("month")
        col_period = field_to_col.get("period")
        col_basic = field_to_col["basic_salary"]

        # Fallbacks for single-payslip files (no id / period columns): infer from
        # the file name, e.g. "payslip-E001-2024-01.xlsx" -> E001, 2024-01.
        fname = os.path.basename(path)
        fn_period = _infer_period_from_name(fname)
        fn_empid = _infer_employee_id_from_name(fname)

        records: list[SalaryRecord] = []
        seen: set[tuple[str, int, int]] = set()

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=FIRST_DATA_ROW), start=FIRST_DATA_ROW
        ):
            cells = {cell.column_letter: cell.value for cell in row}
            basic_raw = cells.get(col_basic)
            id_raw = cells.get(col_id) if col_id else None
            year_raw = cells.get(col_year) if col_year else None
            month_raw = cells.get(col_month) if col_month else None
            period_raw = cells.get(col_period) if col_period else None

            # Fully blank row -> skip (trailing-row tolerance).
            if all(v is None or str(v).strip() == ""
                   for v in (id_raw, year_raw, month_raw, period_raw, basic_raw)):
                continue

            # Employee ID: column value, else inferred from the file name.
            emp_id = str(id_raw).strip() if id_raw not in (None, "") else (fn_empid or "")
            if not emp_id:
                raise ValueError(
                    f"Row {row_idx}: no Employee ID column and couldn't infer one "
                    f"from the file name {fname!r}."
                )

            # Period: Year+Month columns, else a Period/date column, else file name.
            if col_year and col_month and year_raw not in (None, "") and month_raw not in (None, ""):
                year = _parse_year(year_raw, f"{sheet_name}!{col_year}{row_idx}")
                month = _parse_month(month_raw, f"{sheet_name}!{col_month}{row_idx}")
            elif col_period and period_raw not in (None, ""):
                year, month = _parse_period_value(period_raw, f"{sheet_name}!{col_period}{row_idx}")
            elif fn_period:
                year, month = fn_period
            else:
                raise ValueError(
                    f"Row {row_idx}: no Year/Month (or Period) column and couldn't "
                    f"infer the period from the file name {fname!r}."
                )
            if not 1 <= month <= 12:
                raise ValueError(f"Month {month} out of range 1-12 at row {row_idx}.")
            basic = _to_float(
                basic_raw, f"{sheet_name}!{col_basic}{row_idx}", "basic salary"
            )

            key = (emp_id, year, month)
            if key in seen:
                raise ValueError(
                    f"Duplicate payroll period for employee {emp_id!r} "
                    f"{year}-{month:02d} at row {row_idx}; one basic figure "
                    f"per employee-month is expected."
                )
            seen.add(key)

            records.append(
                SalaryRecord(
                    employee_id=emp_id,
                    year=year,
                    month=month,
                    basic_salary=basic,  # basic ONLY — excluded cols ignored
                    source_ref=EvidenceRef(
                        file=path,
                        sheet=sheet_name,
                        cell_or_row=f"{col_basic}{row_idx}",
                        label="basic_salary",
                    ),
                )
            )
        return tuple(records)
    finally:
        wb.close()


def index_by_employee_month(
    records: Tuple[SalaryRecord, ...],
) -> Mapping[tuple[str, int, int], SalaryRecord]:
    """Build an (employee_id, year, month) -> record lookup for G4/G7 (FR-3).

    Parsing already rejects duplicates, so this is a straight projection.
    """
    return {(r.employee_id, r.year, r.month): r for r in records}
