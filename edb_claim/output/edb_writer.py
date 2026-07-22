"""Fill the official EDB RIS(C) output template for submission (FR-5, PLAN.md T10).

Given a per-entity pipeline result, this writes the qualifying personnel into a
copy of ``docs/EDB_Output Template.xlsx`` — the file uploaded to EDB. It is the
*final submission document*.

Audit-fidelity rules (PLAN.md §3 #3 — the easiest place to silently break the
template):

  * write ONLY value cells: ``Manpower_Locals`` cols A–H and the ``Details``
    header values;
  * NEVER touch the col I claim formula (``=ROUND(G*H,2)``) or the hidden
    mapping cols K/L — replicate them for newly-added rows only;
  * keep col K hidden and the row-2 ``=SUM(...)`` totals intact;
  * round only where EDB rounds (the col I formula does it) — write col H at
    full precision so ``ROUND(G*H,2)`` reproduces the cent.

A round-trip fidelity check (:func:`_assert_fidelity`) runs after writing.
"""

from __future__ import annotations

import os
import shutil
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook

from edb_claim.config import Config, settings

# ``Manpower_Locals`` qualifying-category dropdown (col B), read verbatim from
# the template's data validation. We map each free-text designation onto one.
_CATEGORIES = (
    "Directors (only for technical roles)",
    "Manager (only for technical roles)",
    "Engineer",
    "Technician",
    "Scientist / Researcher",
    "Architect / Designer / Developer",
    "Analyst / Consultant",
    "Specialist",
)
_CONFIRM_OK = "Does not breach criteria"  # col C dropdown value for a clean role

_TEMPLATE_DEFAULT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "docs",
    "EDB_Output Template.xlsx",
)
_DATA_START = 5  # Manpower_Locals data rows begin at row 5


def designation_to_category(designation: str) -> str:
    """Best-effort map a designation onto an EDB qualifying category (col B).

    Order matters: an "AI Research Engineer" is an Engineer, not a Researcher.
    """
    d = (designation or "").lower()
    if "technician" in d:
        return "Technician"
    if "engineer" in d or "developer" in d:
        return "Engineer" if "engineer" in d else "Architect / Designer / Developer"
    if "architect" in d or "designer" in d:
        return "Architect / Designer / Developer"
    if "scientist" in d or "research" in d:
        return "Scientist / Researcher"
    if "analyst" in d or "consultant" in d:
        return "Analyst / Consultant"
    if "director" in d:
        return "Directors (only for technical roles)"
    if "manager" in d or "lead" in d or "head" in d:
        return "Manager (only for technical roles)"
    if "specialist" in d:
        return "Specialist"
    return "Engineer"  # neutral default for an RSE


def fill_edb_template(
    entity_result,
    out_path: str,
    *,
    template_path: str = _TEMPLATE_DEFAULT,
    config: Config = settings,
    timestamp: Optional[datetime] = None,
) -> str:
    """Fill the EDB template for ONE entity's qualifying personnel.

    ``entity_result`` is an ``app.pipeline.EntityResult``. Returns ``out_path``.
    """
    shutil.copyfile(template_path, out_path)
    wb = load_workbook(out_path)  # keep formulas (data_only=False default)

    # --- Details header ---------------------------------------------------
    det = wb["Details"]
    det["C3"] = entity_result.entity
    det["C8"] = config.claim_period_start.strftime("%d-%m-%Y")
    det["C9"] = config.claim_period_end.strftime("%d-%m-%Y")
    # Claim type follows the *audit* status, NOT the support-rate confirmation:
    # the system produces an UnauditedClaim until the ACRA Practitioner audits it.
    det["C10"] = "AuditedClaim" if getattr(config, "claim_is_audited", False) else "UnauditedClaim"
    det["C11"] = "No"  # POC: not the final claim
    if timestamp is not None:
        det["C12"] = timestamp.strftime("%d-%m-%Y %H:%M")

    # --- Manpower_Locals rows --------------------------------------------
    ws = wb["Manpower_Locals"]
    qualifying = [e for e in entity_result.employees if e.qualifies]

    row = _DATA_START
    for e in qualifying:
        ws.cell(row, 1, e.employee.name)                                   # (a)
        ws.cell(row, 2, designation_to_category(e.employee.designation))    # (b)
        ws.cell(row, 3, _CONFIRM_OK)                                        # (c)
        if e.monthly_basic_salary is not None:
            ws.cell(row, 4, round(e.monthly_basic_salary, 2))               # (d)
        ef = ws.cell(row, 5, e.involvement_from)                           # (e) From
        et = ws.cell(row, 6, e.involvement_to)                            # (f) To
        ef.number_format = et.number_format = "DD/MM/YYYY"
        ws.cell(row, 7, config.support_rate)                               # (g) % support
        ws.cell(row, 8, e.method_a.qualifying_cost_total)                  # (h) qualifying cost
        _ensure_formulas(ws, row)  # col I/K/L formulas for newly-added rows only
        row += 1

    # never unhide the mapping column K
    ws.column_dimensions["K"].hidden = True

    wb.save(out_path)
    _assert_fidelity(out_path, n_rows=len(qualifying))
    return out_path


def _ensure_formulas(ws, row: int) -> None:
    """Replicate the col I/K/L formulas for a row that has none (new rows).

    Pre-existing formula rows (the template ships rows 5–7) are left untouched.
    """
    i_cell = ws.cell(row, 9)
    if not (isinstance(i_cell.value, str) and i_cell.value.startswith("=")):
        i_cell.value = f"=ROUND(G{row}*H{row},2)"
    k_cell = ws.cell(row, 11)
    if not (isinstance(k_cell.value, str) and k_cell.value.startswith("=")):
        k_cell.value = (
            f'=IF(OR(B{row}="Others", C{row}="Breaches one or more criteria '
            f'(justification to be provided for EDB\'s review)"),1,0)'
        )
    l_cell = ws.cell(row, 12)
    if not (isinstance(l_cell.value, str) and l_cell.value.startswith("=")):
        l_cell.value = (
            f'=IF($C{row}="Breaches one or more criteria '
            f'(justification to be provided for EDB\'s review)",1,0)'
        )


def _assert_fidelity(path: str, *, n_rows: int) -> None:
    """Round-trip check: formulas, hidden col, and totals survived the write."""
    wb = load_workbook(path)  # formulas preserved
    ws = wb["Manpower_Locals"]
    assert ws["H2"].value and str(ws["H2"].value).startswith("=SUM"), "H2 total clobbered"
    assert ws["I2"].value and str(ws["I2"].value).startswith("=SUM"), "I2 total clobbered"
    assert ws.column_dimensions["K"].hidden, "mapping col K was un-hidden"
    last = _DATA_START + max(n_rows, 1) - 1
    iv = ws.cell(last, 9).value
    assert isinstance(iv, str) and iv.startswith("=ROUND"), f"col I formula missing at row {last}: {iv!r}"
