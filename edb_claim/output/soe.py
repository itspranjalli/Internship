"""Statement of Expenditure (SOE) workbook for the public accountant.

The SSRS 4400 auditor samples ≥ 85 % of the claimed value and verifies it
against the Statement of Expenditure. This module renders that pack: the
expenditure summary that ties to the EDB submission, the full month-by-month
*workings* behind every figure, the evidence trail (file/sheet/cell) so the
auditor can sample, the A-vs-B cross-check, and the excluded personnel (never
silently dropped).

Pure presentation: every number comes from the pipeline result; this file
computes nothing. Determinism — pass an explicit ``timestamp`` (no wall clock).
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from edb_claim.config import Config, settings
from edb_claim.output.edb_writer import designation_to_category

# --- light, print-friendly styling -----------------------------------------
_HEAD = Font(bold=True, color="FFFFFF", size=11)
_HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
_TITLE = Font(bold=True, size=14, color="1F4E79")
_SUB = Font(italic=True, size=9, color="595959")
_TOTAL = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0.00"
_PCT = "0.0%"


def _money(x: Optional[float]) -> Optional[float]:
    return None if x is None else round(float(x), 2)


def _header_row(ws, row: int, headers) -> None:
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row, c, text)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _autosize(ws, widths) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_soe(
    result,
    out_path: str,
    *,
    config: Config = settings,
    timestamp: Optional[datetime] = None,
) -> str:
    """Write the SOE workbook for the whole claim. Returns ``out_path``.

    ``result`` is an ``app.pipeline.PipelineResult``.
    """
    wb = Workbook()
    _cover_sheet(wb, result, config, timestamp)
    _soe_sheet(wb, result)
    _workings_sheet(wb, result)
    _crosscheck_sheet(wb, result)
    _evidence_sheet(wb, result)
    _exclusions_sheet(wb, result)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
def _cover_sheet(wb, result, config, timestamp) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Statement of Expenditure (SOE)"
    ws["A1"].font = _TITLE
    ws["A2"] = "EDB Research Incentive Scheme for Companies — RIS(C)"
    ws["A2"].font = _SUB

    rows = [
        ("Application No.", "S26-10249-RIS(C)"),
        ("Claim period", f"{config.claim_period_start:%d %b %Y} – {config.claim_period_end:%d %b %Y}"),
        ("Entities in this claim", ", ".join(e.entity for e in result.entities)),
        ("Support rate applied", f"{result.support_rate:.0%}"
            + ("" if result.support_rate_is_final else "  (ASSUMED — non-final, pending Letter of Award)")),
        ("Qualifying personnel", str(sum(1 for e in result.all_employees if e.qualifies))),
        ("Total claim amount (Method A)", None),  # filled as money below
        ("Prepared", timestamp.strftime("%d %b %Y %H:%M") if timestamp else "—"),
    ]
    r = 4
    for label, val in rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        if label.startswith("Total claim"):
            c = ws.cell(r, 2, _money(result.total_claim_a))
            c.number_format = _MONEY
            c.font = _TOTAL
        else:
            ws.cell(r, 2, val)
        r += 1

    r += 1
    note = ws.cell(r, 1,
        "How to read this pack: the 'Statement of Expenditure' sheet is the "
        "expenditure summary that ties to the EDB submission. 'Workings' shows the "
        "month-by-month calculation behind every figure; 'Evidence' maps each figure "
        "to its source document and cell for sampling; 'Cross-check' compares the EDB "
        "method against the internal method to flag any data anomalies; 'Exclusions' "
        "lists personnel not claimed, with reasons.")
    note.alignment = _WRAP
    note.font = _SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
    if not result.support_rate_is_final:
        w = ws.cell(r + 5, 1,
            "NON-FINAL: the 30% support rate is assumed pending EDB's Letter of Award. "
            "Figures are for preparation and verification, not a final submission.")
        w.font = Font(bold=True, color="C00000")
        ws.merge_cells(start_row=r + 5, start_column=1, end_row=r + 5, end_column=6)
    _autosize(ws, [26, 60, 12, 12, 12, 12])


# ---------------------------------------------------------------------------
def _soe_sheet(wb, result) -> None:
    ws = wb.create_sheet("Statement of Expenditure")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Statement of Expenditure — qualifying manpower (Locals)"
    ws["A1"].font = _TITLE
    headers = ["S/N", "Entity", "Employee ID", "Name", "Qualifying category",
               "Designation", "Monthly basic salary ($)", "Involvement from",
               "Involvement to", "Qualifying cost ($)", "Support %", "Claim amount ($)"]
    _header_row(ws, 3, headers)

    r = 4
    sn = 0
    for ent in result.entities:
        for e in ent.employees:
            if not e.qualifies:
                continue
            sn += 1
            a = e.method_a
            vals = [
                sn, ent.entity, e.employee.id, e.employee.name,
                designation_to_category(e.employee.designation),
                e.employee.designation,
                _money(e.monthly_basic_salary),
                e.involvement_from.strftime("%d/%m/%Y") if e.involvement_from else None,
                e.involvement_to.strftime("%d/%m/%Y") if e.involvement_to else None,
                _money(a.qualifying_cost_total), a.support_rate, _money(a.claim_amount),
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.border = _BORDER
                if c in (7, 10, 12):
                    cell.number_format = _MONEY
                if c == 11:
                    cell.number_format = _PCT
            r += 1

    # total row
    ws.cell(r, 9, "TOTAL").font = _TOTAL
    for c in (10, 12):
        col = get_column_letter(c)
        cell = ws.cell(r, c, f"=SUM({col}4:{col}{r-1})")
        cell.number_format = _MONEY
        cell.font = _TOTAL
        cell.fill = _TOTAL_FILL
    for c in range(1, 13):
        ws.cell(r, c).fill = _TOTAL_FILL
    _autosize(ws, [5, 34, 12, 22, 24, 24, 16, 14, 14, 16, 9, 15])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
def _workings_sheet(wb, result) -> None:
    ws = wb.create_sheet("Workings (Method A)")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Workings — EDB monthly pro-ration (Method A)"
    ws["A1"].font = _TITLE
    ws["A2"] = ("Qualifying cost = Σ  capped salary × month fraction × time contribution. "
                "Capped at $20,000/month. Claim = qualifying cost × support rate (rounded last).")
    ws["A2"].font = _SUB
    headers = ["Employee ID", "Name", "Month", "Capped salary ($)", "Month fraction",
               "Time contribution", "Qualifying cost ($)"]
    _header_row(ws, 4, headers)

    r = 5
    for e in result.all_employees:
        if not e.qualifies:
            continue
        for m in e.method_a.monthly:
            vals = [e.employee.id, e.employee.name, m.month, _money(m.capped_salary),
                    round(m.month_fraction, 6), round(m.time_contribution, 6),
                    _money(m.qualifying_cost)]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.border = _BORDER
                if c in (4, 7):
                    cell.number_format = _MONEY
            r += 1
    _autosize(ws, [12, 22, 8, 16, 14, 16, 18])
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
def _crosscheck_sheet(wb, result) -> None:
    ws = wb.create_sheet("Cross-check")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Cross-check — EDB method (A) vs internal method (B)"
    ws["A1"].font = _TITLE
    ws["A2"] = ("Method A is the submission basis. Method B is the internal hours-ratio "
                "method, run as a data-quality check. A large gap usually signals a data "
                "issue to resolve before audit (e.g. a New Hire with no timesheet).")
    ws["A2"].font = _SUB
    headers = ["Employee ID", "Name", "Method A claim ($)", "Method B claim ($)",
               "Difference ($)", "Difference %", "Status / note"]
    _header_row(ws, 4, headers)

    name = {e.employee.id: e for e in result.all_employees}
    r = 5
    for row in result.variance.rows:
        e = name.get(row.employee_id)
        if row.new_hire_flag:
            note = "⚠ New Hire forced to 100% time, no timesheet — verify before claiming"
        elif row.material:
            note = "⚠ Methods differ materially — check hours vs involvement period"
        else:
            note = "✓ Consistent"
        vals = [row.employee_id, e.employee.name if e else "",
                _money(row.amount_a), _money(row.amount_b), _money(row.delta_abs),
                (round(row.delta_pct, 1) / 100 if row.delta_pct is not None else None), note]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.border = _BORDER
            if c in (3, 4, 5):
                cell.number_format = _MONEY
            if c == 6:
                cell.number_format = _PCT
        r += 1
    _autosize(ws, [12, 22, 18, 18, 16, 12, 52])
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
def _evidence_sheet(wb, result) -> None:
    ws = wb.create_sheet("Evidence")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Evidence trail — every eligibility decision and its source (FR-7)"
    ws["A1"].font = _TITLE
    headers = ["Employee ID", "Name", "Check", "Result", "Reason", "Source file", "Cell / row"]
    _header_row(ws, 3, headers)

    r = 4
    for e in result.all_employees:
        for ev in e.gate_evaluations:
            ref = ev.source_ref
            vals = [e.employee.id, e.employee.name, ev.gate.value,
                    "Pass" if ev.passed else "Fail", ev.reason or "",
                    ref.file if ref else "", ref.cell_or_row if ref else ""]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.border = _BORDER
                if c == 5:
                    cell.alignment = _WRAP
            r += 1
    _autosize(ws, [12, 22, 8, 8, 44, 22, 20])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
def _exclusions_sheet(wb, result) -> None:
    ws = wb.create_sheet("Exclusions")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Personnel not claimed (excluded or blocked) — reported, never dropped"
    ws["A1"].font = _TITLE
    headers = ["Entity", "Employee ID", "Name", "Designation", "Status",
               "Failed checks", "Reason(s)"]
    _header_row(ws, 3, headers)

    r = 4
    for ent in result.entities:
        for e in ent.employees:
            if e.qualifies:
                continue
            vals = [ent.entity, e.employee.id, e.employee.name, e.employee.designation,
                    e.verdict.status.value,
                    ", ".join(g.value for g in e.verdict.failed_gates),
                    "  •  ".join(e.verdict.reasons)]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.border = _BORDER
                if c == 7:
                    cell.alignment = _WRAP
            r += 1
    if r == 4:
        ws.cell(4, 1, "None — every trainee qualifies.").font = _SUB
    _autosize(ws, [34, 12, 22, 24, 12, 16, 56])
    ws.freeze_panes = "A4"
