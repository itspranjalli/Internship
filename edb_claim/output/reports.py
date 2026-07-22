"""HR-facing reports — the action list of who is not being claimed and why.

Separate from the SOE (which is the auditor's pack): this is the workbook HR
uses to *act* — blocked people are fixable (upload the missing document and
re-run); excluded people are not claimable (with the reason to confirm).
"""

from __future__ import annotations

from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_TITLE = Font(bold=True, size=14, color="1F4E79")
_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
_WRAP = Alignment(wrap_text=True, vertical="top")
_BLOCK_FILL = PatternFill("solid", fgColor="FFF2CC")  # amber — fixable
_EXCL_FILL = PatternFill("solid", fgColor="FCE4E4")   # red — not claimable


def _friendly_action(status: str) -> str:
    if status == "BLOCKED":
        return "Fixable — upload the missing document(s) and re-run."
    return "Not claimable — confirm the reason; no action needed unless data is wrong."


def _sheet_headers(ws, row, headers, widths) -> None:
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row, c, text)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_exclusions_report(
    result,
    out_path: str,
    *,
    timestamp: Optional[datetime] = None,
) -> str:
    """Write the 'who is not being claimed' action list. Returns ``out_path``."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues to fix"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Personnel not yet being claimed — action list"
    ws["A1"].font = _TITLE
    ws["A2"] = ("Amber = blocked (fixable: a document is missing). "
                "Red = excluded (not eligible). Nobody is ever silently dropped.")
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    if timestamp:
        ws["A3"] = f"Prepared {timestamp:%d %b %Y %H:%M}"
        ws["A3"].font = Font(size=9, color="595959")

    _sheet_headers(ws, 5,
        ["Entity", "Employee ID", "Name", "Designation", "Status", "Why", "What to do"],
        [32, 12, 22, 24, 12, 50, 40])

    r = 6
    any_rows = False
    for ent in result.entities:
        for e in ent.employees:
            if e.qualifies:
                continue
            any_rows = True
            why = "  •  ".join(e.verdict.reasons) or ", ".join(g.value for g in e.verdict.failed_gates)
            vals = [ent.entity, e.employee.id, e.employee.name, e.employee.designation,
                    e.verdict.status.value, why, _friendly_action(e.verdict.status.value)]
            fill = _BLOCK_FILL if e.verdict.status.value == "BLOCKED" else _EXCL_FILL
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.fill = fill
                if c in (6, 7):
                    cell.alignment = _WRAP
            r += 1
    if not any_rows:
        ws.cell(6, 1, "🎉 Everyone qualifies — nothing to fix.").font = Font(bold=True)
    ws.freeze_panes = "A6"
    wb.save(out_path)
    return out_path
