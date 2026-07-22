"""Evidence-document preview helpers (pure, Streamlit-free, unit-testable).

The audit chatbot cites every figure as ``{file, sheet, cell}`` (FR-7). To let HR
*see* the evidence in its original form before sending it to EDB, the app shows a
preview of the cited worksheet (or PDF) with the exact cell highlighted. This
module holds the non-UI logic so it can be tested without a running app:

  * :func:`resolve_evidence_path` — map a citation's stored filename back to the
    real (uploaded/temp) path on disk, since uploads are saved under temp names.
  * :func:`parse_cell_ref` — turn ``"Time Sheet!G19"`` / ``"I5"`` / a row index
    into ``(sheet, col_index, row)``.
  * :func:`excel_sheet_to_grid` — read a window of an .xlsx worksheet into a plain
    grid (Excel-style column letters + row numbers) with the focus cell located,
    so the UI can render it as a spreadsheet and highlight one cell.

The Streamlit rendering (dialog, download button, PDF embed) lives in
``app/main.py`` and calls these.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

# --- file resolution -------------------------------------------------------


def resolve_evidence_path(file: str, registry: Optional[Dict[str, str]] = None) -> Optional[str]:
    """Resolve a citation ``file`` to a readable path, or None if not found.

    Citations carry the filename the ingest layer recorded (sometimes a basename
    like ``edb_ab12.xlsx``, sometimes the full input path). Uploads are persisted
    under temp names, so the UI keeps a ``registry`` mapping original-name AND
    stored-basename -> real path. Resolution order: the path as-is if it exists,
    then the registry by basename, then by the full string.
    """
    if not file:
        return None
    registry = registry or {}
    if os.path.exists(file):
        return file
    base = os.path.basename(file)
    return registry.get(base) or registry.get(file)


# --- cell-reference parsing ------------------------------------------------

_A1 = re.compile(r"^([A-Za-z]{1,3})(\d+)$")
_ROWONLY = re.compile(r"(?:row\s*)?(\d+)$", re.IGNORECASE)


def _col_to_index(letters: str) -> int:
    """'A'->1, 'B'->2, ... 'Z'->26, 'AA'->27 (1-based, Excel-style)."""
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def _index_to_col(idx: int) -> str:
    """Inverse of :func:`_col_to_index` (1-based)."""
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def parse_cell_ref(cell: Optional[str]) -> Tuple[Optional[str], Optional[int], Optional[int]]:
    """Parse a citation cell into ``(sheet, col_index, row)`` (any may be None).

    Handles ``"Sheet Name!G19"``, ``"G19"``, ``"I5"``, and a bare/row locator
    like ``"19"`` or ``"row 19"``. Unparseable locators yield all-None.
    """
    if not cell:
        return (None, None, None)
    text = str(cell).strip()
    sheet = None
    if "!" in text:
        sheet, text = text.split("!", 1)
        sheet = sheet.strip().strip("'") or None
        text = text.strip()
    m = _A1.match(text)
    if m:
        return (sheet, _col_to_index(m.group(1)), int(m.group(2)))
    m = _ROWONLY.match(text)
    if m:
        return (sheet, None, int(m.group(1)))
    return (sheet, None, None)


# --- worksheet -> grid window ----------------------------------------------


@dataclass
class SheetGrid:
    """A rendered window of a worksheet, ready for the UI to draw as a table."""

    sheet_name: str
    col_letters: List[str] = field(default_factory=list)      # header labels (A, B, ...)
    row_numbers: List[int] = field(default_factory=list)      # Excel row numbers shown
    rows: List[List[str]] = field(default_factory=list)       # stringified cell values
    focus_row: Optional[int] = None                            # Excel row of the cell
    focus_col_letter: Optional[str] = None                     # Excel col of the cell
    truncated: bool = False                                     # window clipped the sheet


def excel_sheet_to_grid(
    path: str,
    sheet: Optional[str] = None,
    *,
    focus_col: Optional[int] = None,
    focus_row: Optional[int] = None,
    row_window: int = 12,
    max_cols: int = 14,
) -> SheetGrid:
    """Read a window of an .xlsx worksheet centred on the focus cell.

    Reads with ``data_only=False`` so stored literals show as-is and formula cells
    show their formula text (transparent for audit). Shows ``row_window`` rows
    either side of ``focus_row`` (or the top of the sheet if unknown) and up to
    ``max_cols`` columns. Raises on a missing/unreadable file (the caller guards).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
        sheet_name = ws.title
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # row window
        if focus_row:
            r0 = max(1, focus_row - row_window)
            r1 = min(max_row, focus_row + row_window)
        else:
            r0, r1 = 1, min(max_row, row_window * 2 + 1)

        # column window: keep the focus column in view, cap the count
        c0 = 1
        c1 = min(max_col, max_cols)
        if focus_col and focus_col > c1:
            c1 = min(max_col, focus_col + 2)
            c0 = max(1, c1 - max_cols + 1)
        truncated = (r1 < max_row) or (c1 - c0 + 1 < max_col)

        col_letters = [_index_to_col(c) for c in range(c0, c1 + 1)]
        row_numbers: List[int] = []
        rows: List[List[str]] = []
        for r in range(r0, r1 + 1):
            row_numbers.append(r)
            line: List[str] = []
            for c in range(c0, c1 + 1):
                val = ws.cell(row=r, column=c).value
                line.append("" if val is None else str(val))
            rows.append(line)

        return SheetGrid(
            sheet_name=sheet_name,
            col_letters=col_letters,
            row_numbers=row_numbers,
            rows=rows,
            focus_row=focus_row,
            focus_col_letter=_index_to_col(focus_col) if focus_col else None,
            truncated=truncated,
        )
    finally:
        wb.close()
