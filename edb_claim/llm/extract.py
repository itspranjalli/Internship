"""FR-9 — LLM document extraction for salary evidence (PLAN.md T16).

Parses a payslip / payroll row (xlsx in the POC; varied-layout PDFs later) into a
fixed JSON schema: ``employee_id``, ``month``, ``basic_salary``, the
allowance/bonus/CPF components **to exclude**, and a payment reference. Every
field carries a confidence (0-1) and a plain-language reason; **low-confidence
results are never discarded or hidden** — they enter flagged and are shown to HR
to confirm or override (FR-9/FR-14). The 0.85 cutoff (``config.confidence_cutoff``)
sets *when* the confirm-prompt is surfaced, not a discard threshold.

HARD BOUNDARY (CLAUDE.md / PRD §7): the LLM **proposes** an extraction; it never
computes the claim. The deterministic ``ingest/salary.py`` path remains the figure
of record. This module's value is twofold: (1) it can read messy documents the
fixed-schema parser cannot, and (2) where both run, it **cross-checks** the
extracted basic against the deterministic figure and surfaces any disagreement for
audit (FR-3 "payslip basic ≠ Staff Costs" family). Degrades gracefully: with no
endpoint, extraction is skipped and the deterministic figures stand.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Mapping, Optional, Tuple

from edb_claim.config import Config, settings
from edb_claim.domain.models import EvidenceRef
from edb_claim.llm.client import LLMClient

_EXTRACT_SCHEMA = {
    "type": "object",
    "required": [
        "employee_id", "month", "basic_salary",
        "excluded_components", "payment_reference",
        "confidence", "confidence_reason",
    ],
    "properties": {
        "employee_id": {"type": "string"},
        "month": {"type": "string"},          # free text; "" if absent
        "basic_salary": {"type": "number"},
        # things to EXCLUDE from qualifying salary (CPF/bonus/AWS/allowance/...)
        "excluded_components": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["name", "amount"],
                "properties": {
                    "name": {"type": "string"},
                    "amount": {"type": "number"},
                },
            },
        },
        "payment_reference": {"type": "string"},
        "confidence": {"type": "number"},
        "confidence_reason": {"type": "string"},
    },
}


@dataclass(frozen=True)
class ExtractedComponent:
    name: str
    amount: float


@dataclass(frozen=True)
class ExtractedPayslip:
    """One LLM-extracted salary record (FR-9/FR-14), nothing discarded.

    ``basic_salary`` is the model's read of basic-only pay; ``excluded`` lists the
    components it identified to keep OUT of the qualifying figure. ``source_ref``
    locates the row (FR-7). ``low_confidence`` drives the HR confirm-prompt.
    """

    employee_id: str
    month: str
    basic_salary: Optional[float]
    excluded: Tuple[ExtractedComponent, ...]
    payment_reference: str
    confidence: Optional[float]
    confidence_reason: str
    source_ref: Optional[EvidenceRef] = None
    used_model: bool = True

    def low_confidence(self, config: Config = settings) -> bool:
        return self.confidence is not None and self.confidence < config.confidence_cutoff


@dataclass(frozen=True)
class ExtractionCrossCheck:
    """Compares an extracted basic against the deterministic figure (audit, FR-3)."""

    employee_id: str
    year: int
    month: int
    deterministic_basic: float
    extracted_basic: Optional[float]
    agrees: bool
    detail: str
    confidence: Optional[float] = None
    confidence_reason: str = ""
    source_ref: Optional[EvidenceRef] = None


def _render_row(row: Mapping[str, Any]) -> str:
    return "\n".join(f"  {k}: {v}" for k, v in row.items() if v not in (None, ""))


def _prompt(row: Mapping[str, Any]) -> str:
    return (
        "Extract salary fields from this payslip/payroll row for an EDB R&D grant. "
        "Qualifying salary is BASIC monthly salary ONLY — CPF, bonus, AWS, "
        "allowances, COLA, overtime, gross/net totals must all be EXCLUDED. Put the "
        "basic figure in basic_salary and every non-basic money item in "
        "excluded_components. If a field is unclear, still return your best value "
        "and lower the confidence with a one-line reason (e.g. 'two basic-pay "
        "candidates', 'amount partly illegible').\n\n"
        "Return JSON: {\"employee_id\": \"...\", \"month\": \"...\", "
        "\"basic_salary\": <number>, \"excluded_components\": [{\"name\": \"...\", "
        "\"amount\": <number>}], \"payment_reference\": \"...\", \"confidence\": "
        "<0..1>, \"confidence_reason\": \"...\"}.\n\n"
        f"ROW:\n{_render_row(row)}"
    )


def extract_row(
    row: Mapping[str, Any],
    *,
    source_ref: Optional[EvidenceRef] = None,
    client: Optional[LLMClient] = None,
    config: Config = settings,
) -> Optional[ExtractedPayslip]:
    """Extract one salary row. Returns ``None`` when no model is available.

    Never raises into the pipeline and never drops a low-confidence result — the
    confidence + reason ride along on the returned object (FR-14).
    """
    use_client = client if client is not None else _safe_client(config)
    if not (use_client is not None and config.llm_enabled):
        return None
    res = use_client.call(_prompt(row), schema=_EXTRACT_SCHEMA, source_ref=
                          ({"file": source_ref.file} if source_ref else None))
    if not (res.ok and res.parsed):
        return None
    p = res.parsed
    comps = tuple(
        ExtractedComponent(str(c.get("name") or ""), _num(c.get("amount")))
        for c in (p.get("excluded_components") or [])
        if isinstance(c, Mapping)
    )
    return ExtractedPayslip(
        employee_id=str(p.get("employee_id") or "").strip(),
        month=str(p.get("month") or "").strip(),
        basic_salary=_opt_num(p.get("basic_salary")),
        excluded=comps,
        payment_reference=str(p.get("payment_reference") or "").strip(),
        confidence=(res.confidence if res.confidence is not None
                    else _opt_num(p.get("confidence"))),
        confidence_reason=str(p.get("confidence_reason") or "").strip(),
        source_ref=source_ref,
    )


def extract_payroll_register(
    path: str,
    *,
    sheet_name: str = "Payroll",
    limit: Optional[int] = None,
    client: Optional[LLMClient] = None,
    config: Config = settings,
) -> List[ExtractedPayslip]:
    """LLM-extract each row of a payroll xlsx (generic read, by header text).

    Reads the sheet generically (header row 1, data from row 2) so it works on
    layouts the fixed-schema parser would reject. ``limit`` caps rows for a quick
    POC pass. Returns ``[]`` when no model is configured.
    """
    use_client = client if client is not None else _safe_client(config)
    if not (use_client is not None and config.llm_enabled):
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out: List[ExtractedPayslip] = []
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=False)
        header_cells = next(rows, None)
        if header_cells is None:
            return []
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        for idx, row in enumerate(rows, start=2):
            values = {headers[i]: c.value for i, c in enumerate(row) if i < len(headers) and headers[i]}
            if all(v in (None, "") for v in values.values()):
                continue
            ref = EvidenceRef(file=path, sheet=sheet_name, cell_or_row=str(idx),
                              label="payslip_row")
            ext = extract_row(values, source_ref=ref, client=use_client, config=config)
            if ext is not None:
                out.append(ext)
            if limit is not None and len(out) >= limit:
                break
    finally:
        wb.close()
    return out


def cross_check(
    extracted: List[ExtractedPayslip],
    salary_records: Tuple[Any, ...],
    *,
    tolerance: float = 0.01,
    config: Config = settings,
) -> List[ExtractionCrossCheck]:
    """Compare extracted basics against the deterministic salary records (FR-3).

    Matches on (employee_id, month) — month from the extracted free-text is parsed
    leniently. Disagreements beyond ``tolerance`` are flagged for HR review; the
    deterministic figure is always the one used downstream (hard boundary).
    """
    det = {(r.employee_id, r.year, r.month): r for r in salary_records}
    checks: List[ExtractionCrossCheck] = []
    for ex in extracted:
        m = _parse_month(ex.month)
        # find a deterministic record by (id, month) across any year present
        rec = next((r for (eid, _y, mm), r in det.items()
                    if eid == ex.employee_id and (m is None or mm == m)), None)
        if rec is None:
            continue
        det_basic = float(rec.basic_salary)
        ext_basic = ex.basic_salary
        agrees = ext_basic is not None and abs(det_basic - ext_basic) <= tolerance
        if ext_basic is None:
            detail = "Model did not return a basic figure; deterministic value stands."
        elif agrees:
            detail = f"Extracted basic ${ext_basic:,.2f} matches the deterministic figure."
        else:
            detail = (f"Extracted basic ${ext_basic:,.2f} differs from the "
                      f"deterministic ${det_basic:,.2f} — review the payslip.")
        checks.append(ExtractionCrossCheck(
            employee_id=rec.employee_id, year=rec.year, month=rec.month,
            deterministic_basic=det_basic, extracted_basic=ext_basic,
            agrees=agrees, detail=detail,
            confidence=ex.confidence, confidence_reason=ex.confidence_reason,
            source_ref=ex.source_ref,
        ))
    return checks


# --- helpers ---------------------------------------------------------------
_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], 1)}


def _parse_month(text: str) -> Optional[int]:
    t = (text or "").strip().lower()
    if not t:
        return None
    if t.isdigit():
        n = int(t)
        return n if 1 <= n <= 12 else None
    for name, num in _MONTHS.items():
        if t.startswith(name):
            return num
    return None


def _num(v: Any) -> float:
    n = _opt_num(v)
    return n if n is not None else 0.0


def _opt_num(v: Any) -> Optional[float]:
    if v is None or isinstance(v, bool):
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def _safe_client(config: Config) -> Optional[LLMClient]:
    try:
        return LLMClient(config)
    except Exception:
        return None
