"""Synthetic test-data generator for the EDB RIS(C) grant POC (PLAN.md T14, PRD §8).

Produces a deterministic ``sample_data/`` fixture set targeting the **Q7 claim
window 2026-01-01 -> 2026-06-30** (PRD §10 Q7, config.claim_period). It writes,
per entity, the documents HR actually produces, matching **exactly** the schemas
the T2/T3 ingest parsers expect:

  * one internal HR workbook per entity  (T2 ``ingest/timesheet.py``):
      - ``Time Sheet``  : header row 18, data from row 19 (cols B..Z)
      - ``Staff Costs`` : header row 14, data from row 15 (cols B..P),
                          paired to the Time Sheet at **offset +4**
                          (Staff Costs row N <-> Time Sheet row N+4).
  * one ECMF ``RSE List`` workbook covering every employee  (T3 ``ingest/rse_list.py``)
  * one ``Payroll`` register (long/tidy, one row per employee-month)  (T3 ``ingest/salary.py``)
  * ``expectations.json`` -- the ground-truth oracle for T21 (DoD §11.3) and T24.

It also embeds all **13 deliberate cases** from PRD §8, labelled per employee in
``expectations.json`` and ``README.md``.

DETERMINISM (PRD §9)
--------------------
No randomness, no ``datetime.now()``. The roster is *constructed* literally
(:data:`ROSTER`), so re-running this script produces value-identical workbooks
every time. (openpyxl embeds a few zip metadata timestamps, so files are not
guaranteed *byte*-identical across runs on different machines; the cell *values*
are identical, which is what the deterministic ingest reads. ``--check`` below
verifies value-identity by re-parsing.)

THE STAFF COSTS DERIVED-COLUMN CHOICE (important)
-------------------------------------------------
In the real workbook the Staff Costs cols J/M/N/O/P ([B]/[D1]/[D2]/[D3]/[E]) are
**formulas**. openpyxl writes formulas but never computes a cached value, and the
T2 parser opens the file ``data_only=True`` -- so it would read ``None`` for any
formula we wrote. Therefore we write the **literal computed values** into those
cells, faithfully applying the documented Staff Costs formulas INCLUDING their
quirks (replicate, do NOT fix -- CLAUDE.md / PRD §6):

  * ``[B] = "N/A" if A < 5000 else MIN(A, 20000)``  (the header calls [B] an
    *annual* figure but the formula yields a *monthly capped* one -- kept as-is).
  * ``[D1] = NETWORKDAYS(C1, C2) * 8.8``
  * ``[D2] = "N/A" if New Hire else <Time Sheet total hours Z>``
  * ``[D3] = 100% if New Hire else D2/D1``   (New-Hire forced 100%, no timesheet)
  * ``[E] = [B] * [D3]``

These literals are what Method B (T7) reconciles against; Method A (T6) ignores
them and recomputes from first principles.

Run::

    .venv/bin/python sample_data/generate.py          # (re)generate the fixtures
    .venv/bin/python sample_data/generate.py --check   # generate + verify ingest
"""

from __future__ import annotations

import calendar
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import date
from typing import Dict, List, Optional, Tuple

# Make the package importable when run as a plain script.
_REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO not in sys.path:
    sys.path.insert(0, _REPO)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from edb_claim.config import settings
from edb_claim.domain.calendar_utils import (
    month_fraction,
    networkdays,
    weekdays_in_month,
)

# --- locked constants (mirror config; re-stated so the file is self-describing)
CLAIM_START = settings.claim_period_start      # date(2026, 1, 1)
CLAIM_END = settings.claim_period_end          # date(2026, 6, 30)
CLAIM_YEAR = CLAIM_START.year                  # 2026
HPD = settings.hours_per_day                   # 8.8
CAP = settings.salary_cap                      # 20000
FLOOR = settings.salary_floor                  # 5000
SUPPORT = settings.support_rate                # 0.30 (ASSUMED, non-final)

# Two participating entities from the canonical list (PRD §8: 2-3 entities).
ENTITY_A = settings.base_entities[1]   # "ST Engineering Advanced Networks & Sensors Pte Ltd"
ENTITY_B = settings.base_entities[2]   # "ST Engineering Digital System Pte Ltd"

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------------
# Roster model -- one Person per employee, carrying everything the three
# document types and the expectations oracle need. Pure data; constructed
# literally below for determinism.
# ---------------------------------------------------------------------------
@dataclass
class Person:
    employee_id: str
    ts_name: str                       # name as written on the Time Sheet
    entity: str
    case_label: str                    # which PRD §8 case this exercises
    # eligibility inputs ----------------------------------------------------
    citizenship: str                   # "Citizen" / "PR" / "Foreigner"
    ecmf_validated: bool
    no_other_grant: bool               # Time Sheet col I (TRUE = compliant w/ G3)
    designation: str
    hire_type: str                     # "New Hire" / "Upskilled" / "Reskilled"
    # involvement / salary --------------------------------------------------
    date_join: date
    date_left: date
    basic_salary: float                # constant monthly basic across the period
    # monthly project hours Jan..Jun (month_number -> hours); months absent = 0
    hours: Dict[int, float] = field(default_factory=dict)
    # document-variation knobs ---------------------------------------------
    rse_name: Optional[str] = None     # name on the RSE list (defaults to ts_name)
    payroll_name: Optional[str] = None # name on the payroll register (defaults to ts_name)
    rse_citizenship: Optional[str] = None  # RSE-list citizenship (defaults to citizenship)
    rse_ecmf: Optional[bool] = None    # RSE-list ECMF flag (defaults to ecmf_validated)
    omit_payroll_months: Tuple[int, ...] = ()   # months with NO payslip row (G7)
    # expectations ----------------------------------------------------------
    expected_verdict: str = "QUALIFIES"
    failed_gate: Optional[str] = None
    expected_reason: str = ""
    note: str = ""

    def name_on(self, doc: str) -> str:
        if doc == "rse":
            return self.rse_name or self.ts_name
        if doc == "payroll":
            return self.payroll_name or self.ts_name
        return self.ts_name


def _full_time_hours(join: date, left: date) -> Dict[int, float]:
    """Full-time monthly project hours over the claim window: weekdays * 8.8.

    For a month only partly inside [join, left] ∩ claim window we still record
    the *full-month* capacity hours -- Method A's ``time_contribution`` clamps to
    1.0, and its ``month_fraction`` is what pro-rates the partial month. (This
    mirrors how HR records a normal full-time month's hours.)
    """
    out: Dict[int, float] = {}
    lo = max(join, CLAIM_START)
    hi = min(left, CLAIM_END)
    for m in range(1, 7):
        last = calendar.monthrange(CLAIM_YEAR, m)[1]
        m_start, m_end = date(CLAIM_YEAR, m, 1), date(CLAIM_YEAR, m, last)
        if m_end < lo or m_start > hi:
            continue
        out[m] = round(weekdays_in_month(CLAIM_YEAR, m) * HPD, 2)
    return out


# ---------------------------------------------------------------------------
# THE ROSTER -- 20 employees over 2 entities, embedding all 13 PRD §8 cases.
# Entity A (rows below) carries the calc-heavy / happy-path cases; Entity B
# carries the exclusion / blocker / LLM-reconciliation cases. Employee IDs are
# stable and globally unique.
# ---------------------------------------------------------------------------
FULL = (date(2026, 1, 1), date(2026, 12, 31))   # involvement spans the whole window


def _build_roster() -> List[Person]:
    A = ENTITY_A
    B = ENTITY_B
    people: List[Person] = []

    # === Entity A =========================================================
    # 1) Standard full-period RSE -- mirrors the EDB worked-example proportions
    #    (full-time, salary < cap). Clean hand-calc oracle.
    people.append(Person(
        employee_id="ANS-001", ts_name="Lim Jia Hao", entity=A,
        case_label="standard_full_period",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="AI Research Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=9500.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="QUALIFIES",
        note="Happy path; both methods agree-ish. salary 9500 < cap.",
    ))
    # 2) Mid-period JOINER -- partial-month pro-ration (Method A). Joins 12 Mar.
    joiner_join = date(2026, 3, 12)
    people.append(Person(
        employee_id="ANS-002", ts_name="Nurul Aisyah Binte Hassan", entity=A,
        case_label="mid_period_joiner",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Machine Learning Engineer", hire_type="Upskilled",
        date_join=joiner_join, date_left=FULL[1], basic_salary=8000.0,
        hours=_full_time_hours(joiner_join, FULL[1]),
        expected_verdict="QUALIFIES",
        note="Joins 2026-03-12 -> Mar fraction 14/22 (non-trivial partial month).",
    ))
    # 3) Mid-period LEAVER -- partial-month pro-ration. Leaves 15 May.
    leaver_left = date(2026, 5, 15)
    people.append(Person(
        employee_id="ANS-003", ts_name="Tan Wei Ming", entity=A,
        case_label="mid_period_leaver+name_variant",
        citizenship="PR", ecmf_validated=True, no_other_grant=True,
        designation="Senior Data Scientist", hire_type="Reskilled",
        date_join=FULL[0], date_left=leaver_left, basic_salary=7000.0,
        hours=_full_time_hours(FULL[0], leaver_left),
        # NAME VARIANT across docs (FR-11): Time Sheet "Tan Wei Ming",
        # RSE list "WEI MING TAN", payroll "Tan, Wei Ming".
        rse_name="WEI MING TAN", payroll_name="Tan, Wei Ming",
        expected_verdict="QUALIFIES",
        note="Leaves 2026-05-15 -> May fraction 11/21. Also the FR-11 name-variant "
             "case: TS 'Tan Wei Ming' / RSE 'WEI MING TAN' / payroll 'Tan, Wei Ming'.",
    ))
    # 4) Salary cap -- 23,000 clamped to 20,000.
    people.append(Person(
        employee_id="ANS-004", ts_name="Rajesh Kumar Pillai", entity=A,
        case_label="salary_cap_23000",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Principal AI Scientist", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=23000.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="QUALIFIES",
        note="Basic 23,000 -> capped at 20,000/month (arithmetic clamp, not a gate).",
    ))
    # 5) New Hire with NO timesheet hours -- Method A vs B divergence + audit flag.
    people.append(Person(
        employee_id="ANS-005", ts_name="Chua Mei Ling", entity=A,
        case_label="new_hire_no_hours",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="AI Engineer", hire_type="New Hire",
        date_join=FULL[0], date_left=FULL[1], basic_salary=6500.0,
        hours={},  # NO project hours recorded anywhere
        expected_verdict="QUALIFIES",
        note="New Hire, zero timesheet hours. Method A qualifying cost = 0 (no "
             "hours), Method B forces [D3]=100% -> B>>A: the variance New-Hire flag.",
    ))
    # 6) Hours > monthly weekday capacity -- cap WARNING (time_contribution clamps to 1).
    over_hours = _full_time_hours(FULL[0], FULL[1])
    over_hours[4] = round(weekdays_in_month(CLAIM_YEAR, 4) * HPD + 60.0, 2)  # April over capacity
    people.append(Person(
        employee_id="ANS-006", ts_name="Goh Boon Keng", entity=A,
        case_label="hours_over_capacity",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Computer Vision Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=10000.0,
        hours=over_hours,
        expected_verdict="QUALIFIES",
        note="April hours exceed weekday capacity (22*8.8 + 60) -> WARNING; "
             "time_contribution clamps to 1.0, so the claim is unaffected.",
    ))
    # 7) Ambiguous designation -- FR-10 borderline G5 judgement -> HR review queue.
    people.append(Person(
        employee_id="ANS-007", ts_name="Simon Tay Chin Hock", entity=A,
        case_label="ambiguous_designation",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Engineering Operations Manager", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=12000.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        # Deterministic pipeline cannot resolve borderline G5 -> treated as a
        # review item; for the ORACLE we record the expected human resolution
        # (qualifies: it is an engineering role, not Marketing/HR/etc.).
        expected_verdict="QUALIFIES",
        note="Designation 'Engineering Operations Manager' is borderline for G5 "
             "(FR-10) -> LLM/HR review queue. Expected resolution: QUALIFIES "
             "(engineering, not a non-qualifying category).",
    ))

    # === Entity B (exclusions / blocker) ==================================
    # 8) Salary 4,800 -- G4 floor -> EXCLUDED.
    people.append(Person(
        employee_id="DSG-001", ts_name="Faridah Binte Omar", entity=B,
        case_label="salary_below_floor_4800",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Junior AI Engineer", hire_type="New Hire",
        date_join=FULL[0], date_left=FULL[1], basic_salary=4800.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="EXCLUDED", failed_gate="G4",
        expected_reason="Basic monthly salary 4,800 < floor 5,000 (G4).",
    ))
    # 9) Foreigner -- G1 -> EXCLUDED.
    people.append(Person(
        employee_id="DSG-002", ts_name="Arjun Mehta", entity=B,
        case_label="foreigner",
        citizenship="Foreigner", ecmf_validated=True, no_other_grant=True,
        designation="AI Research Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=11000.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="EXCLUDED", failed_gate="G1",
        expected_reason="Not a local (Foreigner) -> fails G1; flagged, not claimed.",
    ))
    # 10) Not ECMF-validated -- G2 -> EXCLUDED.
    people.append(Person(
        employee_id="DSG-003", ts_name="Wong Kah Wai", entity=B,
        case_label="not_ecmf_validated",
        citizenship="Citizen", ecmf_validated=False, no_other_grant=True,
        designation="Data Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=9000.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        rse_ecmf=False,  # RSE list agrees: not validated
        expected_verdict="EXCLUDED", failed_gate="G2",
        expected_reason="Not ECMF-validated (Time Sheet + RSE list) -> fails G2.",
    ))
    # 11) Enjoying another government grant -- G3 -> EXCLUDED.
    people.append(Person(
        employee_id="DSG-004", ts_name="Priya Nair", entity=B,
        case_label="other_grant",
        citizenship="PR", ecmf_validated=True, no_other_grant=False,
        designation="NLP Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=10500.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="EXCLUDED", failed_gate="G3",
        expected_reason="Enjoying another government cash grant (col I) -> fails G3.",
    ))
    # 12) Designation "HR Manager" -- G5 non-qualifying -> EXCLUDED.
    people.append(Person(
        employee_id="DSG-005", ts_name="Kelvin Ong Wei Sheng", entity=B,
        case_label="non_qualifying_designation_hr",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="HR Manager", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=9500.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="EXCLUDED", failed_gate="G5",
        expected_reason="Designation 'HR Manager' is in the non-qualifying set (HR) -> fails G5.",
    ))
    # 13) Missing payslip for one month -- G7 BLOCKER + re-upload demo.
    blk = _full_time_hours(FULL[0], FULL[1])
    people.append(Person(
        employee_id="DSG-006", ts_name="Siti Nurhaliza Rahman", entity=B,
        case_label="missing_payslip_one_month",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="MLOps Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=8800.0,
        hours=blk,
        omit_payroll_months=(4,),  # April payslip missing
        expected_verdict="BLOCKED", failed_gate="G7",
        expected_reason="No payslip evidence for 2026-04 (G7 blocker on that "
                        "person-month) -> claim blocked pending re-upload.",
    ))

    # --- a few extra plain QUALIFIES rows to round each entity to ~10 ------
    people.append(Person(
        employee_id="ANS-008", ts_name="Lee Hui Shan", entity=A,
        case_label="standard_full_period_extra",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="AI Research Scientist", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=14000.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="QUALIFIES",
        note="Additional standard qualifying RSE (entity-A headcount filler).",
    ))
    people.append(Person(
        employee_id="ANS-009", ts_name="Mohammed Iqbal Bin Yusof", entity=A,
        case_label="standard_full_period_extra",
        citizenship="PR", ecmf_validated=True, no_other_grant=True,
        designation="Robotics Engineer", hire_type="Reskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=8500.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="QUALIFIES",
        note="Additional standard qualifying RSE.",
    ))
    people.append(Person(
        employee_id="ANS-010", ts_name="Cheryl Teo Xin Yi", entity=A,
        case_label="standard_full_period_extra",
        citizenship="Citizen", ecmf_validated=True, no_other_grant=True,
        designation="Deep Learning Engineer", hire_type="Upskilled",
        date_join=FULL[0], date_left=FULL[1], basic_salary=11500.0,
        hours=_full_time_hours(FULL[0], FULL[1]),
        expected_verdict="QUALIFIES",
        note="Additional standard qualifying RSE.",
    ))
    for i, (nm, sal, des, cit, ht) in enumerate([
        ("Daniel Koh Jun Wei", 9200.0, "AI Platform Engineer", "Citizen", "Upskilled"),
        ("Anitha Devi Krishnan", 10800.0, "Data Scientist", "PR", "Upskilled"),
        ("Bryan Sim Hong Yi", 7800.0, "AI Software Engineer", "Citizen", "Reskilled"),
        ("Ng Pei Shan", 13500.0, "Senior ML Engineer", "Citizen", "Upskilled"),
    ], start=7):
        people.append(Person(
            employee_id=f"DSG-{i:03d}", ts_name=nm, entity=B,
            case_label="standard_full_period_extra",
            citizenship=cit, ecmf_validated=True, no_other_grant=True,
            designation=des, hire_type=ht,
            date_join=FULL[0], date_left=FULL[1], basic_salary=sal,
            hours=_full_time_hours(FULL[0], FULL[1]),
            expected_verdict="QUALIFIES",
            note="Additional standard qualifying RSE (entity-B headcount filler).",
        ))

    return people


ROSTER = _build_roster()


# ---------------------------------------------------------------------------
# Method A hand-calc oracle (reuses calendar_utils -- PRD §6 formula).
# Full precision throughout; round ONLY the final claim like EDB's
# I = ROUND(G*H, 2). This is the cent-level control for T21 / §11.3.
# ---------------------------------------------------------------------------
def method_a_handcalc(p: Person) -> dict:
    """Compute Method A qualifying cost + claim for ``p`` over the claim window.

    Returns the per-month breakdown and the totals. Matches PRD §6:
      qualifying_cost = Σ capped_salary * month_fraction(m) * time_contribution(m)
      time_contribution(m) = min(1, hours(m) / (weekdays(m) * 8.8))
      claim_amount = round(qualifying_cost * support_rate, 2)
    """
    capped = min(p.basic_salary, CAP)
    lo = max(p.date_join, CLAIM_START)
    hi = min(p.date_left, CLAIM_END)
    months = []
    total = 0.0
    for m in range(1, 7):
        mf = month_fraction(CLAIM_YEAR, m, lo, hi)
        if mf == 0.0:
            continue
        wd = weekdays_in_month(CLAIM_YEAR, m)
        hrs = float(p.hours.get(m, 0.0))
        capacity = wd * HPD
        tc = min(1.0, hrs / capacity) if capacity else 0.0
        qc = capped * mf * tc
        total += qc
        months.append({
            "month": m, "weekdays": wd, "hours": hrs,
            "month_fraction": mf, "time_contribution": tc,
            "capped_salary": capped, "qualifying_cost": qc,
        })
    claim = round(total * SUPPORT, 2)
    return {
        "capped_salary": capped,
        "qualifying_cost_total": total,
        "support_rate": SUPPORT,
        "claim_amount": claim,
        "monthly": months,
    }


def method_b_values(p: Person) -> dict:
    """Replicate the Staff Costs [B]/[D1]/[D2]/[D3]/[E] formulas (quirks intact).

    Used both to fill the Staff Costs literal cells AND to record Method B in the
    expectations file. ``[D2]`` = full-window total project hours (sum of the
    Time Sheet months), ``[D1]`` = NETWORKDAYS(join,left)*8.8. New Hire forces
    [D2]="N/A" and [D3]=100%. [B] is "N/A" below floor, else MIN(salary,cap).
    """
    new_hire = (p.hire_type == "New Hire")
    b = "N/A" if p.basic_salary < FLOOR else float(min(p.basic_salary, CAP))
    d1 = networkdays(p.date_join, p.date_left) * HPD
    total_hours = round(sum(p.hours.values()), 2)
    d2 = "N/A" if new_hire else total_hours
    if new_hire:
        d3 = 1.0  # forced 100% (quirk, replicated)
    else:
        d3 = (d2 / d1) if d1 else 0.0
    if b == "N/A":
        e = "N/A"
    else:
        e = b * d3
    return {
        "new_hire": new_hire,
        "B": b, "D1": d1, "D2": d2, "D3": d3, "E": e,
    }


# ---------------------------------------------------------------------------
# Workbook writers (match the T2/T3 layouts exactly).
# ---------------------------------------------------------------------------
def _write_internal_workbook(entity: str, people: List[Person], path: str) -> None:
    """Write one internal HR workbook (Time Sheet + Staff Costs) for ``entity``.

    Time Sheet : header row 18, data row 19+ (cols B..Z; entity name at G3).
    Staff Costs: header row 14, data row 15+; row N pairs Time Sheet row N+4.
    Derived Staff Costs cols are written as LITERAL VALUES (see module docstring).
    """
    wb = Workbook()
    ts = wb.active
    ts.title = "Time Sheet"

    # --- Time Sheet headers (match ingest/timesheet.py _TS_WANTED) ---------
    ts["B3"] = "To be completed on a monthly basis"
    ts["G3"] = entity                      # TS_ENTITY_CELL
    ts["B17"] = "PROJECT TEAM"
    ts_headers = {
        "B": "S/N", "C": "Employee ID", "D": "Name", "E": "Local/ Foreigner",
        "F": "Academic qualifications", "G": "Employee designation",
        "H": "Validated by ECMF",
        "I": "Confirm not enjoying any other government grant",
        "J": "Current AI Proficiency Level", "K": "AI Capabilities*",
        "L": "New Hirer/Upskilled/Reskilled",
        "M": "Date of Completion for Upskilling or Reskilling",
        "N": "Jan", "O": "Feb", "P": "Mar", "Q": "Apr", "R": "May", "S": "Jun",
        "T": "Jul", "U": "Aug", "V": "Sep", "W": "Oct", "X": "Nov", "Y": "Dec",
        "Z": "Total",
    }
    for col, txt in ts_headers.items():
        ts[f"{col}18"] = txt

    month_cols = {1: "N", 2: "O", 3: "P", 4: "Q", 5: "R", 6: "S"}  # Jan..Jun

    # --- Time Sheet data rows (from 19) ------------------------------------
    for i, p in enumerate(people):
        r = 19 + i
        ts[f"B{r}"] = i + 1
        ts[f"C{r}"] = p.employee_id
        ts[f"D{r}"] = p.ts_name
        ts[f"E{r}"] = p.citizenship          # "Citizen"/"PR"/"Foreigner"
        ts[f"F{r}"] = "Bachelor of Engineering"
        ts[f"G{r}"] = p.designation
        ts[f"H{r}"] = bool(p.ecmf_validated)  # TRUE/FALSE bool
        ts[f"I{r}"] = bool(p.no_other_grant)  # TRUE = compliant with G3
        ts[f"J{r}"] = "PL3"
        ts[f"K{r}"] = "Machine Learning"
        ts[f"L{r}"] = p.hire_type
        if p.hire_type in ("Upskilled", "Reskilled"):
            ts[f"M{r}"] = date(2026, 1, 31)
        total = 0.0
        for m, col in month_cols.items():
            h = p.hours.get(m)
            if h is not None and h != 0:
                ts[f"{col}{r}"] = h
                total += h
        # Total Z written as a literal (data_only ingest reads values).
        ts[f"Z{r}"] = round(total, 2)

    # --- Staff Costs sheet -------------------------------------------------
    sc = wb.create_sheet("Staff Costs")
    sc["B13"] = "HUMAN RESOURCES"
    sc["I13"] = "Actual monthly salary per IR8A"
    sc["J13"] = "Total qualifying annual salary1"
    sc["K13"] = "Date join"
    sc["L13"] = "Date left"
    sc["M13"] = "Total hours in employment (not adjusted for leave)"
    sc["N13"] = "Time spent on project (Hours)"
    sc["O13"] = "% of time spent on Project "
    sc["P13"] = "Staff Cost for RISC Claims"
    sc_headers = {
        "B": "S/N", "C": "Employee ID", "D": "Name", "E": "Local/ Foreigner",
        "F": "Academic qualifications", "G": "Designations",
        "H": "New Hirer/Upskilled/Reskilled",
        "I": "[A]", "J": "[B]", "K": "[C1]", "L": "[C2]",
        "M": "[D1] = [C2 - C1] * 8.8 Hrs", "N": "[D2]",
        "O": "[D3] = [D2/D1]", "P": "[E] = [B * D3]",
    }
    for col, txt in sc_headers.items():
        sc[f"{col}14"] = txt

    # Staff Costs row N pairs Time Sheet row N+4 (offset +4). The Time Sheet
    # data starts at row 19, so the first Staff Costs data row is 15.
    for i, p in enumerate(people):
        r = 15 + i
        mb = method_b_values(p)
        sc[f"B{r}"] = i + 1
        sc[f"C{r}"] = p.employee_id          # literal Employee ID (cross-ref value)
        sc[f"D{r}"] = p.ts_name
        sc[f"E{r}"] = p.citizenship
        sc[f"F{r}"] = "Bachelor of Engineering"
        sc[f"G{r}"] = p.designation
        sc[f"H{r}"] = p.hire_type
        sc[f"I{r}"] = float(p.basic_salary)  # [A]
        sc[f"J{r}"] = mb["B"]                 # [B] number or "N/A"
        sc[f"K{r}"] = p.date_join             # [C1]
        sc[f"L{r}"] = p.date_left             # [C2]
        sc[f"M{r}"] = mb["D1"]                # [D1]
        sc[f"N{r}"] = mb["D2"]                # [D2] number or "N/A"
        sc[f"O{r}"] = mb["D3"]                # [D3]
        sc[f"P{r}"] = mb["E"]                 # [E] number or "N/A"

    wb.save(path)


def _write_rse_list(people: List[Person], path: str) -> None:
    """Write the ECMF ``RSE List`` workbook (T3 ingest/rse_list.py schema).

    Header row 1: A Employee ID | B Name | C Citizenship | D ECMF Validated.
    Covers EVERY employee across all entities. Citizenship/ECMF reflect the
    per-case truth (e.g. foreigner, non-ECMF). Name may differ from the Time
    Sheet for the FR-11 name-variant case.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "RSE List"
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = (
        "Employee ID", "Name", "Citizenship", "ECMF Validated",
    )
    r = 2
    for p in people:
        ws[f"A{r}"] = p.employee_id
        ws[f"B{r}"] = p.name_on("rse")
        ws[f"C{r}"] = p.rse_citizenship or p.citizenship
        ecmf = p.rse_ecmf if p.rse_ecmf is not None else p.ecmf_validated
        ws[f"D{r}"] = "TRUE" if ecmf else "FALSE"
        r += 1
    wb.save(path)


def _write_payroll(people: List[Person], path: str) -> None:
    """Write the ``Payroll`` register (T3 ingest/salary.py schema, long format).

    Header row 1 includes the four required columns PLUS noise columns
    (Allowances/Bonus/AWS/CPF/Gross) to exercise basic-only isolation. One row
    per employee-month over Jan..Jun, EXCEPT months in ``omit_payroll_months``
    (the G7 missing-payslip case) and months outside [join, left].
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    headers = [
        "Employee ID", "Name", "Year", "Month", "Basic Salary",
        "Allowances", "Bonus", "AWS", "CPF (Employer)", "CPF (Employee)", "Gross Pay",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)

    r = 2
    for p in people:
        for m in range(1, 7):
            # only months the person is employed within the window
            last = calendar.monthrange(CLAIM_YEAR, m)[1]
            m_start, m_end = date(CLAIM_YEAR, m, 1), date(CLAIM_YEAR, m, last)
            if m_end < p.date_join or m_start > p.date_left:
                continue
            if m in p.omit_payroll_months:
                continue  # G7: deliberately no payslip for this month
            basic = float(p.basic_salary)
            ws.cell(r, 1, p.employee_id)
            ws.cell(r, 2, p.name_on("payroll"))
            ws.cell(r, 3, CLAIM_YEAR)
            ws.cell(r, 4, m)
            ws.cell(r, 5, basic)
            # noise columns -- never read by the basic-only parser
            ws.cell(r, 6, round(basic * 0.10, 2))   # Allowances
            ws.cell(r, 7, round(basic * 0.50, 2) if m == 6 else 0.0)  # Bonus (mid-year)
            ws.cell(r, 8, round(basic, 2) if m == 6 else 0.0)         # AWS
            ws.cell(r, 9, round(basic * 0.17, 2))   # CPF employer
            ws.cell(r, 10, round(basic * 0.20, 2))  # CPF employee
            ws.cell(r, 11, round(basic * 1.30, 2))  # Gross (aggregate)
            r += 1
    wb.save(path)


# ---------------------------------------------------------------------------
# Expectations (ground-truth oracle).
# ---------------------------------------------------------------------------
# Which employees get a full Method-A hand-calc to the cent (PRD §11.3 requires
# >= 3, including the partial-month joiner).
_HANDCALC_IDS = ("ANS-001", "ANS-002", "ANS-003", "ANS-004", "ANS-005")


def _build_expectations() -> dict:
    employees = []
    for p in ROSTER:
        rec = {
            "employee_id": p.employee_id,
            "entity": p.entity,
            "ts_name": p.ts_name,
            "case_label": p.case_label,
            "expected_verdict": p.expected_verdict,
            "failed_gate": p.failed_gate,
            "expected_reason": p.expected_reason,
            "note": p.note,
        }
        # Method A hand-calc for the selected qualifying employees, to the cent.
        if p.employee_id in _HANDCALC_IDS:
            a = method_a_handcalc(p)
            rec["method_a"] = {
                "qualifying_cost_total": round(a["qualifying_cost_total"], 6),
                "support_rate": a["support_rate"],
                "claim_amount": a["claim_amount"],  # the cent-level oracle
                "monthly": [
                    {
                        "month": mo["month"],
                        "weekdays": mo["weekdays"],
                        "hours": mo["hours"],
                        "month_fraction": round(mo["month_fraction"], 6),
                        "time_contribution": round(mo["time_contribution"], 6),
                        "capped_salary": mo["capped_salary"],
                        "qualifying_cost": round(mo["qualifying_cost"], 6),
                    }
                    for mo in a["monthly"]
                ],
            }
            # Method B values (for variance reconciliation / T8).
            b = method_b_values(p)
            e = b["E"]
            rec["method_b"] = {
                "B": b["B"], "D1": round(b["D1"], 6),
                "D2": b["D2"], "D3": round(b["D3"], 6),
                "E": (round(e, 6) if isinstance(e, (int, float)) else e),
                "claim_amount": (
                    round(e * SUPPORT, 2) if isinstance(e, (int, float)) else None
                ),
                "new_hire": b["new_hire"],
            }
        employees.append(rec)

    return {
        "_meta": {
            "description": "Ground-truth oracle for the EDB RIS(C) POC synthetic "
                           "data (PRD §8 cases). Generated by sample_data/generate.py.",
            "claim_window": [CLAIM_START.isoformat(), CLAIM_END.isoformat()],
            "support_rate": SUPPORT,
            "support_rate_is_final": settings.support_rate_is_final,
            "salary_floor": FLOOR,
            "salary_cap": CAP,
            "hours_per_day": HPD,
            "entities": [ENTITY_A, ENTITY_B],
            "method_a_handcalc_ids": list(_HANDCALC_IDS),
            "cent_oracle_note": (
                "Method A: claim = round(Σ capped*month_fraction*time_contribution "
                "* support_rate, 2). Round ONLY the final claim (PRD §6). "
                "The 7,310.87 control lives in the EDB template, not here."
            ),
            "files": {
                "internal_workbook_entity_a": "internal_ANS.xlsx",
                "internal_workbook_entity_b": "internal_DSG.xlsx",
                "rse_list": "rse_list.xlsx",
                "payroll": "payroll.xlsx",
            },
        },
        "employees": employees,
    }


# ---------------------------------------------------------------------------
# Orchestration.
# ---------------------------------------------------------------------------
ENTITY_FILES = {
    ENTITY_A: "internal_ANS.xlsx",
    ENTITY_B: "internal_DSG.xlsx",
}


def generate() -> dict:
    """Generate all fixtures into ``sample_data/`` and return the manifest dict."""
    by_entity: Dict[str, List[Person]] = {ENTITY_A: [], ENTITY_B: []}
    for p in ROSTER:
        by_entity[p.entity].append(p)

    for entity, fname in ENTITY_FILES.items():
        _write_internal_workbook(entity, by_entity[entity], os.path.join(OUT_DIR, fname))

    _write_rse_list(ROSTER, os.path.join(OUT_DIR, "rse_list.xlsx"))
    _write_payroll(ROSTER, os.path.join(OUT_DIR, "payroll.xlsx"))

    expectations = _build_expectations()
    with open(os.path.join(OUT_DIR, "expectations.json"), "w") as fh:
        json.dump(expectations, fh, indent=2, sort_keys=False)
        fh.write("\n")

    return expectations


def main(argv: List[str]) -> int:
    generate()
    print(f"Generated fixtures in {OUT_DIR}:")
    for f in sorted(ENTITY_FILES.values()) + ["rse_list.xlsx", "payroll.xlsx", "expectations.json"]:
        print(f"  - {f}")
    if "--check" in argv:
        from sample_data._verify import verify
        return verify()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
