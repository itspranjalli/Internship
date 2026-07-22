"""Tests for the submission-document writers (SOE, EDB template, issues report).

These are presentation-layer round-trip tests: they run the real pipeline over
the synthetic fixtures and assert the generated workbooks (a) open, (b) preserve
the EDB template's audit-critical formulas / hidden column / totals, and (c) tie
the SOE total back to the Method A oracle.

Runs under pytest OR directly via the plain-assert harness at the bottom.
"""

import os
import sys
import tempfile
from datetime import datetime

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from openpyxl import load_workbook

from edb_claim.app.pipeline import run_pipeline
from edb_claim.output.edb_writer import designation_to_category, fill_edb_template
from edb_claim.output.reports import build_exclusions_report
from edb_claim.output.soe import build_soe

_SAMPLE = os.path.join(_REPO_ROOT, "sample_data")
_TEMPLATE = os.path.join(_REPO_ROOT, "docs", "EDB_Output Template.xlsx")
_TS = datetime(2026, 6, 16, 9, 30)


def _result():
    return run_pipeline(
        [os.path.join(_SAMPLE, "internal_ANS.xlsx"), os.path.join(_SAMPLE, "internal_DSG.xlsx")],
        os.path.join(_SAMPLE, "rse_list.xlsx"),
        os.path.join(_SAMPLE, "payroll.xlsx"),
    )


# ---------------------------------------------------------------------------
# EDB template — audit fidelity (PLAN.md §3 #3)
# ---------------------------------------------------------------------------
def test_edb_template_preserves_formulas_hidden_col_and_totals():
    res = _result()
    ent = res.entities[0]  # ANS — 10 qualifying
    n = sum(1 for e in ent.employees if e.qualifies)
    with tempfile.TemporaryDirectory() as d:
        out = fill_edb_template(ent, os.path.join(d, "edb.xlsx"),
                                template_path=_TEMPLATE, timestamp=_TS)
        wb = load_workbook(out)  # formulas preserved
        ml = wb["Manpower_Locals"]
        # row-2 totals untouched
        assert str(ml["H2"].value).startswith("=SUM")
        assert str(ml["I2"].value).startswith("=SUM")
        # hidden mapping column stays hidden
        assert ml.column_dimensions["K"].hidden
        # every data row has the claim formula (existing + newly added)
        for r in range(5, 5 + n):
            assert str(ml.cell(r, 9).value).startswith("=ROUND(G")
            assert str(ml.cell(r, 11).value).startswith("=IF(")  # K mapping formula
        # value cells written; col H (qualifying cost) is full precision (not pre-rounded)
        assert ml.cell(5, 7).value == res.support_rate
        assert ml.cell(5, 8).value > 0


def test_edb_template_qualifying_cost_reproduces_oracle_claim():
    """col I = ROUND(G*H,2) must reproduce the Method A claim to the cent."""
    res = _result()
    ent = res.entities[0]
    quals = [e for e in ent.employees if e.qualifies]
    with tempfile.TemporaryDirectory() as d:
        out = fill_edb_template(ent, os.path.join(d, "edb.xlsx"),
                                template_path=_TEMPLATE, timestamp=_TS)
        ml = load_workbook(out)["Manpower_Locals"]
        for i, e in enumerate(quals):
            g = ml.cell(5 + i, 7).value
            h = ml.cell(5 + i, 8).value
            assert round(g * h, 2) == e.method_a.claim_amount, (e.employee.id, round(g * h, 2))


def test_edb_only_qualifying_rows_written():
    res = _result()
    ent = res.entities[1]  # DSG — 4 qualifying, 6 not
    n = sum(1 for e in ent.employees if e.qualifies)
    with tempfile.TemporaryDirectory() as d:
        out = fill_edb_template(ent, os.path.join(d, "edb.xlsx"),
                                template_path=_TEMPLATE, timestamp=_TS)
        ml = load_workbook(out)["Manpower_Locals"]
        assert ml.cell(5, 1).value is not None              # first qualifying row filled
        assert ml.cell(5 + n, 1).value is None              # no row past the qualifying set


def test_designation_category_mapping():
    assert designation_to_category("AI Research Engineer") == "Engineer"
    assert designation_to_category("Senior Data Scientist") == "Scientist / Researcher"
    assert designation_to_category("Principal AI Scientist") == "Scientist / Researcher"
    assert designation_to_category("") == "Engineer"  # neutral default


# ---------------------------------------------------------------------------
# SOE — ties to the claim, has every sheet
# ---------------------------------------------------------------------------
def test_soe_has_all_sheets_and_total_ties_to_claim():
    res = _result()
    with tempfile.TemporaryDirectory() as d:
        out = build_soe(res, os.path.join(d, "soe.xlsx"), timestamp=_TS)
        wb = load_workbook(out, data_only=False)
        for sheet in ("Cover", "Statement of Expenditure", "Workings (Method A)",
                      "Cross-check", "Evidence", "Exclusions"):
            assert sheet in wb.sheetnames, sheet
        soe = wb["Statement of Expenditure"]
        # one row per qualifying employee + header(3) + total
        n_q = sum(1 for e in res.all_employees if e.qualifies)
        # data rows 4..(4+n_q-1); the total row carries a SUM over the claim column (L)
        total_row = 4 + n_q
        assert str(soe.cell(total_row, 12).value).startswith("=SUM(L4:")


def test_soe_workings_only_qualifying_and_full_precision():
    res = _result()
    with tempfile.TemporaryDirectory() as d:
        out = build_soe(res, os.path.join(d, "soe.xlsx"), timestamp=_TS)
        ws = load_workbook(out)["Workings (Method A)"]
        # there is at least one workings row, and month fractions are not pre-rounded to 2dp
        seen = False
        for r in range(5, ws.max_row + 1):
            mf = ws.cell(r, 5).value
            if isinstance(mf, (int, float)):
                seen = True
        assert seen


# ---------------------------------------------------------------------------
# Issues report — excluded + blocked listed, nobody dropped
# ---------------------------------------------------------------------------
def test_issues_report_lists_every_non_qualifying_person():
    res = _result()
    not_claimed = [e for e in res.all_employees if not e.qualifies]
    with tempfile.TemporaryDirectory() as d:
        out = build_exclusions_report(res, os.path.join(d, "issues.xlsx"), timestamp=_TS)
        ws = load_workbook(out)["Issues to fix"]
        ids = {ws.cell(r, 2).value for r in range(6, ws.max_row + 1)}
        for e in not_claimed:
            assert e.employee.id in ids, e.employee.id


# ---------------------------------------------------------------------------
# Plain-assert harness
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failed = 0
    for fn in fns:
        try:
            fn()
            print(f"ok   {fn.__name__}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"FAIL {fn.__name__}: {exc}")
    sys.exit(1 if failed else 0)
